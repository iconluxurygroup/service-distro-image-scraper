import logging
import pandas as pd
import pyodbc
import asyncio
import json
import datetime
import os
import time
import hashlib
import psutil
import httpx
from fastapi import BackgroundTasks
from sqlalchemy.sql import text
from sqlalchemy.exc import SQLAlchemyError
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from database_config import conn_str, async_engine
from s3_utils import upload_file_to_space
from db_utils import (
    get_send_to_email,
    get_images_excel_db,
    update_file_location_complete,
    update_file_generate_complete,
    fetch_last_valid_entry,
    update_log_url_in_db,
)
from search_utils import update_search_sort_order, insert_search_results
from common import fetch_brand_rules
from utils import create_temp_dirs, cleanup_temp_dirs, generate_search_variations, process_and_tag_results
from endpoint_utils import sync_get_endpoint
from logging_config import setup_job_logger
from email_utils import send_message_email
import aiofiles
import aiohttp
from typing import Optional, List, Dict, Tuple
from urllib.parse import urlparse
from url_extract import extract_thumbnail_url
import re
from operator import itemgetter
from PIL import Image as IMG2
from io import BytesIO
import numpy as np
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

default_logger = logging.getLogger(__name__)
if not default_logger.handlers:
    default_logger.setLevel(logging.INFO)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

BRAND_RULES_URL = os.getenv("BRAND_RULES_URL", "https://raw.githubusercontent.com/iconluxurygroup/legacy-icon-product-api/refs/heads/main/task_settings/brand_settings.json")

# Image Processing Functions (Optimized for Async Smoothness)
async def download_all_images(
    image_list: List[Dict],
    temp_dir: str,
    logger: Optional[logging.Logger] = None,
    batch_size: int = 10
) -> List[Tuple[str, int]]:
    logger = logger or default_logger
    failed_downloads = []
    logger.info(f"📥 Starting download of {len(image_list)} images to {temp_dir}")

    if not image_list:
        logger.error("Image list is empty")
        return []

    os.makedirs(temp_dir, exist_ok=True)
    logger.info(f"📁 Ensured directory exists: {temp_dir}")

    def clean_url(url: str, attempt: int = 1) -> str:
        try:
            if not url:
                return url
            if attempt == 1:
                url = re.sub(r'\\+|%5[Cc]', '', url)
                parsed = urlparse(url)
                path = parsed.path.replace('%2F', '/').replace('%2f', '/')
                cleaned_url = f"{parsed.scheme}://{parsed.netloc}{path}"
                if parsed.query:
                    query = parsed.query.replace('%5C', '').replace('%5c', '')
                    cleaned_url += f"?{query}"
                if parsed.fragment:
                    cleaned_url += f"#{parsed.fragment}"
                return cleaned_url
            elif attempt == 2:
                url = re.sub(r'\\+|%5[Cc]|%2[Ff]', '', url)
                parsed = urlparse(url)
                path = parsed.path
                cleaned_url = f"{parsed.scheme}://{parsed.netloc}{path}"
                if parsed.query:
                    cleaned_url += f"?{parsed.query}"
                if parsed.fragment:
                    cleaned_url += f"#{parsed.fragment}"
                return cleaned_url
            elif attempt == 3:
                url = re.sub(r'[\x00-\x1F\x7F]', '', url)
                parsed = urlparse(url)
                return f"{parsed.scheme}://{parsed.netloc}{parsed.path}" + \
                       (f"?{parsed.query}" if parsed.query else "") + \
                       (f"#{parsed.fragment}" if parsed.fragment else "")
            return url
        except Exception as e:
            logger.warning(f"Error cleaning URL {url} on attempt {attempt}: {e}")
            return url

    async def validate_url(url: str, session: aiohttp.ClientSession, logger: logging.Logger) -> bool:
        try:
            if not url or not re.match(r'^https?://', url):
                logger.warning(f"Invalid URL format: {url}")
                return False
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/91.0.4472.124',
                'Accept': 'image/*,*/*;q=0.8',
                'Referer': 'https://www.google.com/'
            }
            async with session.head(url, timeout=5, headers=headers, allow_redirects=True) as response:
                if response.status == 200:
                    logger.debug(f"URL {url} is accessible")
                    return True
                elif response.status == 404:
                    logger.warning(f"URL {url} is permanently unavailable (404)")
                    return False
                logger.warning(f"URL {url} returned status {response.status}")
                return False
        except aiohttp.ClientError as e:
            logger.warning(f"URL {url} is not accessible: {e}")
            return False
        except asyncio.TimeoutError:
            logger.warning(f"Timeout validating URL {url}")
            return False

    @retry(
        stop=stop_after_attempt(lambda attempt, kwargs: 4 if kwargs.get('entry_index', 0) >= 2 else 3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        retry=retry_if_exception_type((aiohttp.ClientResponseError, asyncio.TimeoutError)),
        before_sleep=lambda retry_state: retry_state.kwargs['logger'].info(
            f"Retrying download for URL {retry_state.kwargs['url']} (attempt {retry_state.attempt_number}/{4 if retry_state.kwargs.get('entry_index', 0) >= 2 else 3}) after {retry_state.next_action.sleep}s"
        )
    )
    async def download_image(
        url: str,
        filename: str,
        session: aiohttp.ClientSession,
        logger: logging.Logger,
        entry_index: int = 0,
        timeout: int = 30,
        max_clean_attempts: int = 3
    ) -> bool:
        if not url:
            logger.error(f"Empty URL for filename {filename}")
            return False
        extracted_url = extract_thumbnail_url(url, logger)
        logger.debug(f"Extracted URL: {extracted_url}")
        for attempt in range(1, max_clean_attempts + 1):
            try:
                cleaned_url = clean_url(extracted_url, attempt)
                logger.debug(f"Attempt {attempt} - Raw URL: {url}, Cleaned URL: {cleaned_url}")
                if not await validate_url(cleaned_url, session, logger):
                    logger.warning(f"Attempt {attempt} - Skipping inaccessible URL: {cleaned_url}")
                    continue
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/91.0.4472.124',
                    'Accept': 'image/*,*/*;q=0.8',
                    'Referer': 'https://www.google.com/'
                }
                async with session.get(cleaned_url, timeout=timeout, headers=headers, allow_redirects=True) as response:
                    if response.status != 200:
                        logger.warning(f"Attempt {attempt} - HTTP error for image {cleaned_url}: {response.status} {response.reason}")
                        if response.status == 404:
                            return False
                        continue
                    async with aiofiles.open(filename, 'wb') as f:
                        content = await response.read()
                        if not content:
                            logger.error(f"Empty content downloaded for {cleaned_url}")
                            return False
                        await f.write(content)
                    if not os.path.exists(filename) or os.path.getsize(filename) == 0:
                        logger.error(f"Downloaded file {filename} is missing or empty")
                        return False
                    logger.debug(f"Attempt {attempt} - Successfully downloaded {cleaned_url} to {filename}")
                    return True
            except aiohttp.ClientResponseError as e:
                if e.status == 404:
                    logger.error(f"Attempt {attempt} - Permanent failure for {url}: 404 Not Found")
                    return False
                logger.error(f"Attempt {attempt} - HTTP error for image {url}: {str(e)}")
                raise
            except asyncio.TimeoutError:
                logger.error(f"Attempt {attempt} - Timeout downloading image {url}")
                raise
            except Exception as e:
                logger.error(f"Attempt {attempt} - Error downloading image {url}: {str(e)}", exc_info=True)
                continue
        logger.error(f"All {max_clean_attempts} attempts failed for URL {url}")
        return False

    sort_strategies = [
        lambda lst: sorted(lst, key=lambda x: 1 if x.get('ImageUrlThumbnail') else 0, reverse=True),
        lambda lst: sorted(lst, key=lambda x: len(x.get('ImageUrl', '')), reverse=False),
        lambda lst: sorted(lst, key=lambda x: urlparse(x.get('ImageUrl', '')).netloc or ''),
        lambda lst: sorted(lst, key=lambda x: x.get('ExcelRowID', 0)),
    ]

    async def process_image(image: Dict, index: int, semaphore: asyncio.Semaphore) -> None:
        async with semaphore:
            excel_row_id = image.get('ExcelRowID')
            main_url = image.get('ImageUrl')
            thumb_url = image.get('ImageUrlThumbnail', '')
            if not excel_row_id or not isinstance(excel_row_id, (int, str)):
                logger.error(f"Invalid ExcelRowID: {excel_row_id}")
                failed_downloads.append(("Invalid ExcelRowID", excel_row_id))
                return
            filename = os.path.join(temp_dir, f"{excel_row_id}_image.jpg")

            logger.debug(f"Processing ExcelRowID {excel_row_id} at index {index}: Main URL = {main_url}, Thumbnail URL = {thumb_url}")

            async with aiohttp.ClientSession() as session:
                success = await download_image(main_url, filename, session, logger, entry_index=index)
                if not success and thumb_url:
                    logger.debug(f"Falling back to thumbnail {thumb_url} for ExcelRowID {excel_row_id}")
                    success = await download_image(thumb_url, filename, session, logger, entry_index=index)
                
                if not success:
                    logger.error(f"Failed to download both main and thumbnail for ExcelRowID {excel_row_id}")
                    failed_downloads.append((main_url or thumb_url or "No valid URL", excel_row_id))
                    if not main_url and not thumb_url:
                        logger.critical(f"No valid URLs for ExcelRowID {excel_row_id}. Check database for FileID {image.get('FileID', 'unknown')}.")

    best_failed_downloads = []
    min_failures = float('inf')
    original_tail = image_list[2:] if len(image_list) > 2 else []
    fixed_prefix = image_list[:2] if len(image_list) >= 2 else image_list
    semaphore = asyncio.Semaphore(batch_size)  # Control concurrency

    for strategy_idx, sort_func in enumerate(sort_strategies, 1):
        failed_downloads.clear()
        logger.info(f"Trying sort strategy {strategy_idx}")
        sorted_tail = sort_func(original_tail)
        current_list = fixed_prefix + sorted_tail
        logger.debug(f"Current list order: {[item.get('ExcelRowID', 'N/A') for item in current_list]}")

        tasks = [process_image(image, idx, semaphore) for idx, image in enumerate(current_list)]
        await asyncio.gather(*tasks, return_exceptions=True)
        logger.info(f"Sort strategy {strategy_idx} completed with {len(failed_downloads)} failures")

        if len(failed_downloads) < min_failures:
            min_failures = len(failed_downloads)
            best_failed_downloads = failed_downloads.copy()
        if min_failures == 0:
            break

    if best_failed_downloads:
        logger.info(f"Retrying {len(best_failed_downloads)} failed downloads with fallback strategy")
        retry_failed = []
        async with aiohttp.ClientSession() as session:
            for url, excel_row_id in best_failed_downloads:
                image = next((img for img in image_list if img.get('ExcelRowID') == excel_row_id), None)
                if not image:
                    logger.warning(f"No image data found for ExcelRowID {excel_row_id} during retry")
                    retry_failed.append((url, excel_row_id))
                    continue
                filename = os.path.join(temp_dir, f"{excel_row_id}_image.jpg")
                main_url = clean_url(image.get('ImageUrl', ''), attempt=3)
                success = await download_image(main_url, filename, session, logger, entry_index=0)
                if not success and image.get('ImageUrlThumbnail'):
                    thumb_url = clean_url(image.get('ImageUrlThumbnail', ''), attempt=3)
                    success = await download_image(thumb_url, filename, session, logger, entry_index=0)
                if not success:
                    retry_failed.append((url, excel_row_id))
        best_failed_downloads = retry_failed

    logger.info(f"📸 Completed image downloads. Total failed: {len(best_failed_downloads)}/{len(image_list)}")
    return best_failed_downloads

async def verify_png_image_single(image_path: str, logger: Optional[logging.Logger] = None) -> bool:
    logger = logger or default_logger
    try:
        if not os.path.exists(image_path):
            logger.error(f"Image file does not exist: {image_path}")
            return False
        logger.debug(f"🔎 Verifying image: {image_path}")
        async with aiofiles.open(image_path, 'rb') as f:
            img_data = await f.read()
        if not img_data:
            logger.error(f"Empty image data for {image_path}")
            return False
        img = await asyncio.to_thread(IMG2.open, BytesIO(img_data))
        if img is None:
            logger.error(f"❌ Failed to open image: {image_path}")
            return False
        await asyncio.to_thread(img.verify)
        logger.info(f"✅ Image verified successfully: {image_path}")

        image_size = (await aiofiles.os.stat(image_path)).st_size
        logger.debug(f"📏 Image size: {image_size} bytes")

        if image_size < 3000:
            logger.warning(f"⚠️ File may be corrupted or too small: {image_path}")
            return False

        if not await resize_image(image_path, logger=logger):
            logger.warning(f"⚠️ Resize failed for: {image_path}")
            return False

        async with aiofiles.open(image_path, 'rb') as f:
            img_data = await f.read()
        if not img_data:
            logger.error(f"Empty image data after resize for {image_path}")
            return False
        img = await asyncio.to_thread(IMG2.open, BytesIO(img_data))
        if img is None:
            logger.error(f"❌ Failed to open image after resize: {image_path}")
            return False
        await asyncio.to_thread(img.verify)
        logger.info(f"✅ Post-resize verification successful: {image_path}")
        return True
    except Exception as e:
        logger.error(f"❌ Image verify failed: {e}, for image: {image_path}", exc_info=True)
        return False

async def resize_image(image_path: str, logger: Optional[logging.Logger] = None) -> bool:
    logger = logger or default_logger
    try:
        logger.debug(f"📂 Attempting to open image: {image_path}")
        async with aiofiles.open(image_path, 'rb') as f:
            img_data = await f.read()
        if not img_data:
            logger.error(f"Empty image data for {image_path}")
            return False
        img = await asyncio.to_thread(IMG2.open, BytesIO(img_data))
        logger.debug(f"After opening: img={img}, type={type(img)}, has_size={hasattr(img, 'size')}")
        if img is None or not hasattr(img, 'size'):
            logger.error(f"❌ Invalid image object after opening: {image_path}")
            return False
        
        MAXSIZE = 130

        if img.mode == 'RGBA':
            logger.info(f"🌈 Converting RGBA image to RGB with white background: {image_path}")
            background = await asyncio.to_thread(IMG2.new, 'RGB', img.size, (255, 255, 255))
            await asyncio.to_thread(background.paste, img, mask=img.split()[3])
            img = background
        elif img.mode != 'RGB':
            logger.info(f"🌈 Converting {img.mode} image to RGB: {image_path}")
            img = await asyncio.to_thread(img.convert, 'RGB')
        logger.debug(f"After mode conversion: img={img}, type={type(img)}, has_size={hasattr(img, 'size')}")
        if not hasattr(img, 'size'):
            logger.error(f"❌ Image object is invalid after mode conversion: {image_path}")
            return False

        def get_background_color(img):
            logger.debug(f"Inside get_background_color: img={img}, type={type(img)}")
            width, height = img.size
            pixels = np.array(img)
            top = pixels[0, :]
            bottom = pixels[height-1, :]
            left = pixels[1:height-1, 0]
            right = pixels[1:height-1, width-1]
            border_pixels = np.concatenate((top, bottom, left, right))
            color_counts = Counter(map(tuple, border_pixels))
            most_common_color, count = color_counts.most_common(1)[0]
            total_border_pixels = border_pixels.shape[0]
            if count / total_border_pixels >= 0.9:
                return most_common_color
            return None

        def is_white(color, threshold=240):
            r, g, b = color
            return r >= threshold and g >= threshold and b >= threshold

        def replace_background(img, bg_color):
            pixels = np.array(img)
            bg_color = np.array(bg_color)
            white = np.array([255, 255, 255])
            diff = np.abs(pixels - bg_color)
            mask = np.all(diff <= 5, axis=2)
            pixels[mask] = white
            return IMG2.fromarray(pixels)

        logger.debug(f"Before get_background_color: img={img}, type={type(img)}")
        background_color = await asyncio.to_thread(get_background_color, img)
        if background_color and not await asyncio.to_thread(is_white, background_color):
            logger.info(f"🖌️ Replacing background color {background_color} with white for {image_path}")
            img = await asyncio.to_thread(replace_background, img, background_color)

        h, w = img.height, img.width
        logger.debug(f"📐 Original size: height={h}, width={w}")
        if h > MAXSIZE or w > MAXSIZE:
            if h > w:
                w = int(w * MAXSIZE / h)
                h = MAXSIZE
            else:
                h = int(h * MAXSIZE / w)
                w = MAXSIZE
            logger.debug(f"🔍 Resizing to: height={h}, width={w}")
            new_img = await asyncio.to_thread(img.resize, (w, h))
        else:
            new_img = img

        buffer = BytesIO()
        await asyncio.to_thread(new_img.save, buffer, format='PNG')
        async with aiofiles.open(image_path, 'wb') as f:
            content = buffer.getvalue()
            if not content:
                logger.error(f"Empty buffer after saving image {image_path}")
                return False
            await f.write(content)
        logger.info(f"✅ Image processed and saved: {image_path}")

        if await aiofiles.os.path.exists(image_path):
            file_size = (await aiofiles.os.stat(image_path)).st_size
            logger.debug(f"📏 File size after save: {file_size} bytes")
            if file_size == 0:
                logger.error(f"Resized file {image_path} is empty")
                return False
        else:
            logger.error(f"❌ File not found after save: {image_path}")
            return False
        return True
    except Exception as e:
        logger.error(f"❌ Error resizing image: {e}, for image: {image_path}", exc_info=True)
        return False

async def anchor_images_to_excel(
    excel_file: str,
    temp_dir: str,
    image_data: List[Dict],
    column: str = "A",
    row_offset: int = 5,
    preferred_image_method: str = "append",
    logger: Optional[logging.Logger] = None
) -> List[int]:
    logger = logger or default_logger
    failed_rows = []

    try:
        if not os.path.exists(excel_file):
            logger.warning(f"Excel file {excel_file} does not exist, creating new")
        logger.debug(f"📂 Loading or creating workbook at {excel_file}")
        if os.path.exists(excel_file):
            wb = await asyncio.to_thread(load_workbook, excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Images"
            headers = ["Image", "Brand", "Style", "Color", "Category", "Status", "Width", "Height", "Format"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header

        if ws.max_row < row_offset:
            for row in range(2, row_offset + 1):
                ws.cell(row=row, column=1).value = ""

        logger.info(f"🖼️ Processing images for {excel_file} with method {preferred_image_method}")
        if not await aiofiles.os.path.exists(temp_dir):
            logger.error(f"❌ Temp directory does not exist: {temp_dir}")
            return failed_rows

        image_files = await asyncio.to_thread(os.listdir, temp_dir)
        logger.debug(f"Found {len(image_files)} files in temp_dir: {image_files}")
        image_map = {}
        for f in image_files:
            if '_' in f and f.split('_')[0].isdigit():
                row_id = int(f.split('_')[0])
                image_map[row_id] = f
            else:
                logger.warning(f"Skipping invalid filename {f} in temp_dir")

        for item in image_data:
            row_id = item.get('ExcelRowID')
            if not row_id or not isinstance(row_id, (int, str)):
                logger.error(f"Invalid ExcelRowID: {row_id}")
                failed_rows.append(row_id)
                continue
            try:
                row_id_int = int(row_id)
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid row_id type for {row_id}: expected int, got {type(row_id)}, error: {e}")
                failed_rows.append(row_id)
                continue
            row_number = row_id_int + row_offset
            logger.debug(f"Processing row_id={row_id_int}, row_number={row_number}")

            image_path = os.path.join(temp_dir, image_map.get(row_id_int, ""))
            status = "Failed"
            width, height, format_ = 0, 0, "Unknown"

            if row_id_int in image_map:
                if await verify_png_image_single(image_path, logger):
                    img = await asyncio.to_thread(OpenpyxlImage, image_path)
                    img.width, img.height = 80, 80
                    if preferred_image_method in ["append", "overwrite"]:
                        anchor = f"{column}{row_number}"
                    elif preferred_image_method == "NewColumn":
                        anchor = f"B{row_number}"
                    else:
                        logger.error(f"Unrecognized preferred image method: {preferred_image_method}")
                        failed_rows.append(row_id_int)
                        continue

                    try:
                        await asyncio.to_thread(ws.add_image, img, anchor)
                        ws.row_dimensions[row_number].height = 60
                        logger.info(f"✅ Image added at {anchor}")
                    except Exception as e:
                        logger.error(f"Failed to anchor image at {anchor}: {e}", exc_info=True)
                        failed_rows.append(row_id_int)
                        continue

                    async with aiofiles.open(image_path, 'rb') as f:
                        img_data = await f.read()
                    if not img_data:
                        logger.error(f"Empty image data for {image_path} after anchoring")
                        failed_rows.append(row_id_int)
                        continue
                    pil_img = await asyncio.to_thread(IMG2.open, BytesIO(img_data))
                    width, height, format_ = pil_img.width, pil_img.height, pil_img.format or "Unknown"
                    status = "Success"
                else:
                    logger.warning(f"⚠️ Image verification failed for {image_path}")
                    failed_rows.append(row_id_int)
            else:
                logger.warning(f"⚠️ No image found for row {row_id_int}")
                failed_rows.append(row_id_int)

            ws[f"B{row_number}"] = item.get('Brand', '')
            ws[f"C{row_number}"] = item.get('Style', '')
            ws[f"D{row_number}"] = item.get('Color', '')
            ws[f"E{row_number}"] = item.get('Category', '')
            ws[f"F{row_number}"] = status
            ws[f"G{row_number}"] = width
            ws[f"H{row_number}"] = height
            ws[f"I{row_number}"] = format_

            if not ws[f"B{row_number}"].value:
                logger.warning(f"⚠️ Missing Brand in B{row_number}")
            if not ws[f"C{row_number}"].value:
                logger.warning(f"⚠️ Missing Style in C{row_number}")

        max_data_row = max(int(item.get('ExcelRowID', 0)) for item in image_data) + row_offset if image_data else row_offset
        if ws.max_row > max_data_row:
            logger.info(f"🗑️ Removing rows {max_data_row + 1} to {ws.max_row}")
            await asyncio.to_thread(ws.delete_rows, max_data_row + 1, ws.max_row - max_data_row)

        try:
            await asyncio.to_thread(wb.save, excel_file)
            file_size = os.path.getsize(excel_file) if os.path.exists(excel_file) else 0
            logger.info(f"🏁 Excel file saved with anchored images: {excel_file}, size: {file_size} bytes")
            if file_size < 1000:
                logger.warning(f"Excel file {excel_file} is suspiciously small ({file_size} bytes)")
        except Exception as e:
            logger.error(f"Failed to save Excel file {excel_file}: {e}", exc_info=True)
            return failed_rows

        return failed_rows
    except Exception as e:
        logger.error(f"❌ Error anchoring images to Excel: {e}", exc_info=True)
        return failed_rows

async def write_failed_downloads_to_excel(
    failed_downloads: List[Tuple[str, int]],
    excel_file: str,
    logger: Optional[logging.Logger] = None
) -> bool:
    logger = logger or default_logger
    if failed_downloads:
        try:
            logger.debug(f"📂 Loading workbook from {excel_file}")
            if not os.path.exists(excel_file):
                logger.error(f"Excel file {excel_file} does not exist for writing failed downloads")
                return False
            wb = await asyncio.to_thread(load_workbook, excel_file)
            ws = wb.create_sheet("FailedDownloads") if "FailedDownloads" not in wb else wb["FailedDownloads"]

            max_row = ws.max_row or 1
            max_failed_row = max(row_id for _, row_id in failed_downloads) if failed_downloads else max_row
            if max_failed_row > max_row:
                logger.debug(f"Extending worksheet from {max_row} to {max_failed_row} rows")
                for i in range(max_row + 1, max_failed_row + 1):
                    ws[f"A{i}"] = ""

            for url, row_id in failed_downloads:
                cell_reference = f"A{row_id}"
                logger.debug(f"✍️ Writing URL {url} to cell {cell_reference}")
                ws[cell_reference] = str(url) if url else "No valid URL"
                ws[cell_reference].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            await asyncio.to_thread(wb.save, excel_file)
            logger.info(f"✅ Failed downloads written to Excel file: {excel_file}")
            return True
        except Exception as e:
            logger.error(f"❌ Error writing failed downloads to Excel: {e}", exc_info=True)
            return False
    logger.info(f"ℹ️ No failed downloads to write to Excel.")
    return True

async def process_images_and_anchor(
    image_list: List[Dict],
    temp_dir: str,
    excel_file: str,
    preferred_image_method: str = "append",
    logger: Optional[logging.Logger] = None
) -> bool:
    logger = logger or default_logger
    logger.info(f"🚀 Starting image processing pipeline with method {preferred_image_method}")

    if not image_list:
        logger.error("Image list is empty, cannot process")
        return False

    if preferred_image_method not in ["append", "overwrite", "NewColumn"]:
        logger.warning(f"Invalid preferred_image_method '{preferred_image_method}', defaulting to 'append'")
        preferred_image_method = "append"

    try:
        failed_downloads = await download_all_images(image_list, temp_dir, logger)
        logger.debug(f"Download completed with {len(failed_downloads)} failed downloads")
        
        failed_rows = await anchor_images_to_excel(
            excel_file, temp_dir, image_data=image_list, 
            preferred_image_method=preferred_image_method, logger=logger
        )
        logger.debug(f"Anchoring completed with {len(failed_rows)} failed rows")
        
        success = await write_failed_downloads_to_excel(failed_downloads, excel_file, logger)
        logger.debug(f"Writing failed downloads completed with success={success}")
        
        logger.info(f"🎉 Pipeline completed. Failed downloads: {len(failed_downloads)}, Failed rows: {len(failed_rows)}")
        return success and len(failed_rows) == 0
    except Exception as e:
        logger.error(f"Pipeline failed: {e}", exc_info=True)
        return False

# Original process_restart_batch (As Provided)
async def process_restart_batch(
    file_id_db: int,
    entry_id: Optional[int] = None,
    use_all_variations: bool = False,
    logger: Optional[logging.Logger] = None
) -> Dict[str, str]:
    log_filename = f"job_logs/job_{file_id_db}.log"
    try:
        if logger is None:
            logger, log_filename = setup_job_logger(job_id=str(file_id_db), log_dir="job_logs", console_output=True)
        logger.setLevel(logging.DEBUG)
        process = psutil.Process()
        logger.debug(f"Logger initialized")

        def log_memory_usage():
            mem_info = process.memory_info()
            logger.info(f"Memory: RSS={mem_info.rss / 1024**2:.2f} MB")
            if mem_info.rss / 1024**2 > 1000:
                logger.warning(f"High memory usage")

        logger.info(f"Starting processing for FileID: {file_id_db}")
        log_memory_usage()

        file_id_db_int = file_id_db
        BATCH_SIZE = 1
        MAX_CONCURRENCY = 4

        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = :file_id"),
                {"file_id": file_id_db_int}
            )
            if result.fetchone()[0] == 0:
                logger.error(f"FileID {file_id_db} does not exist")
                return {"error": f"FileID {file_id_db} does not exist", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}
            result.close()

        if entry_id is None:
            entry_id = await fetch_last_valid_entry(str(file_id_db_int), logger)
            if entry_id is not None:
                async with async_engine.connect() as conn:
                    result = await conn.execute(
                        text("SELECT MIN(EntryID) FROM utb_ImageScraperRecords WHERE FileID = :file_id AND EntryID > :entry_id"),
                        {"file_id": file_id_db_int, "entry_id": entry_id}
                    )
                    next_entry = result.fetchone()
                    entry_id = next_entry[0] if next_entry and next_entry[0] else None
                    logger.info(f"Resuming from EntryID: {entry_id}")
                    result.close()

        brand_rules = await fetch_brand_rules(BRAND_RULES_URL, max_attempts=3, timeout=10, logger=logger)
        if not brand_rules:
            logger.warning(f"No brand rules fetched")
            return {"message": "Failed to fetch brand rules", "file_id": str(file_id_db), "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        endpoint = None
        for attempt in range(5):
            try:
                endpoint = sync_get_endpoint(logger=logger)
                if endpoint:
                    logger.info(f"Selected endpoint: {endpoint}")
                    break
                logger.warning(f"Attempt {attempt + 1} failed")
                await asyncio.sleep(2)
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed: {e}")
                await asyncio.sleep(2)
        if not endpoint:
            logger.error(f"No healthy endpoint")
            return {"error": "No healthy endpoint", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        async with async_engine.connect() as conn:
            query = text("""
                SELECT EntryID, ProductModel, ProductBrand, ProductColor, ProductCategory 
                FROM utb_ImageScraperRecords 
                WHERE FileID = :file_id AND (:entry_id IS NULL OR EntryID >= :entry_id) 
                ORDER BY EntryID
            """)
            result = await conn.execute(query, {"file_id": file_id_db_int, "entry_id": entry_id})
            entries = [(row[0], row[1], row[2], row[3], row[4]) for row in result.fetchall() if row[1] is not None]
            logger.info(f"Found {len(entries)} entries")
            result.close()

        if not entries:
            logger.warning(f"No valid EntryIDs found for FileID {file_id_db}")
            return {"error": "No entries found", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        entry_batches = [entries[i:i + BATCH_SIZE] for i in range(0, len(entries), BATCH_SIZE)]
        logger.info(f"Created {len(entry_batches)} batches")

        successful_entries = 0
        failed_entries = 0
        last_entry_id_processed = entry_id or 0
        required_columns = ["EntryID", "ImageUrl", "ImageDesc", "ImageSource", "ImageUrlThumbnail"]

        semaphore = asyncio.Semaphore(MAX_CONCURRENCY)
        async def process_entry(entry):
            entry_id, search_string, brand, color, category = entry
            async with semaphore:
                try:
                    logger.info(f"Processing EntryID {entry_id}")
                    results = await async_process_entry_search(
                        search_string=search_string,
                        brand=brand,
                        endpoint=endpoint,
                        entry_id=entry_id,
                        use_all_variations=use_all_variations,
                        file_id_db=file_id_db,
                        logger=logger
                    )
                    if not results:
                        logger.error(f"No results for EntryID {entry_id}")
                        return entry_id, False

                    if not all(all(col in res for col in required_columns) for res in results):
                        logger.error(f"Missing columns for EntryID {entry_id}")
                        return entry_id, False

                    deduplicated_results = []
                    seen = set()
                    for res in results:
                        key = (res['EntryID'], res['ImageUrl'])
                        if key not in seen:
                            seen.add(key)
                            deduplicated_results.append(res)
                    logger.info(f"Deduplicated to {len(deduplicated_results)} rows")

                    insert_success = await insert_search_results(deduplicated_results, logger=logger, file_id=str(file_id_db))
                    if not insert_success:
                        logger.error(f"Failed to insert results for EntryID {entry_id}")
                        return entry_id, False

                    update_result = await update_search_sort_order(
                        str(file_id_db), str(entry_id), brand, search_string, color, category, logger, brand_rules=brand_rules
                    )
                    if update_result is None or not update_result:
                        logger.error(f"SortOrder update failed for EntryID {entry_id}")
                        return entry_id, False

                    return entry_id, True
                except Exception as e:
                    logger.error(f"Error processing EntryID {entry_id}: {e}", exc_info=True)
                    return entry_id, False

        for batch_idx, batch_entries in enumerate(entry_batches, 1):
            logger.info(f"Processing batch {batch_idx}/{len(entry_batches)}")
            start_time = datetime.datetime.now()

            results = await asyncio.gather(
                *(process_entry(entry) for entry in batch_entries),
                return_exceptions=True
            )

            for entry, result in zip(batch_entries, results):
                entry_id = entry[0]
                if isinstance(result, Exception):
                    logger.error(f"Error processing EntryID {entry_id}: {result}", exc_info=True)
                    failed_entries += 1
                    continue
                entry_id_result, success = result
                if success:
                    successful_entries += 1
                    last_entry_id_processed = entry_id
                else:
                    failed_entries += 1

            elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
            logger.info(f"Completed batch {batch_idx} in {elapsed_time:.2f}s")
            log_memory_usage()

        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("""
                    SELECT COUNT(DISTINCT t.EntryID), 
                           SUM(CASE WHEN t.SortOrder > 0 THEN 1 ELSE 0 END) AS positive_count,
                           SUM(CASE WHEN t.SortOrder IS NULL THEN 1 ELSE 0 END) AS null_count
                    FROM utb_ImageScraperResult t
                    INNER JOIN utb_ImageScraperRecords r ON t.EntryID = r.EntryID
                    WHERE r.FileID = :file_id
                """),
                {"file_id": file_id_db_int}
            )
            row = result.fetchone()
            total_entries = row[0] if row else 0
            positive_entries = row[1] if row and row[1] is not None else 0
            null_entries = row[2] if row and row[2] is not None else 0
            logger.info(
                f"Verification: {total_entries} total entries, "
                f"{positive_entries} with positive SortOrder, {null_entries} with NULL SortOrder"
            )
            if null_entries > 0:
                logger.warning(f"Found {null_entries} entries with NULL SortOrder")
            result.close()

        to_emails = await get_send_to_email(file_id_db, logger=logger)
        if to_emails:
            subject = f"Processing Completed for FileID: {file_id_db}"
            message = (
                f"Processing for FileID {file_id_db} completed.\n"
                f"Successful entries: {successful_entries}/{len(entries)}\n"
                f"Failed entries: {failed_entries}\n"
                f"Last EntryID: {last_entry_id_processed}\n"
                f"Log file: {log_filename}"
            )
            await send_message_email(to_emails, subject=subject, message=message, logger=logger)

        log_public_url = await upload_log_file(str(file_id_db), log_filename, logger)
        return {
            "message": "Search processing completed",
            "file_id": str(file_id_db),
            "successful_entries": str(successful_entries),
            "total_entries": str(len(entries)),
            "failed_entries": str(failed_entries),
            "log_filename": log_filename,
            "log_public_url": log_public_url or "",
            "last_entry_id": str(last_entry_id_processed)
        }
    except Exception as e:
        logger.error(f"Error processing FileID {file_id_db}: {e}", exc_info=True)
        log_public_url = await upload_log_file(str(file_id_db), log_filename, logger)
        return {"error": str(e), "log_filename": log_filename, "log_public_url": log_public_url or "", "last_entry_id": str(entry_id or "")}
    finally:
        await async_engine.dispose()
        logger.info(f"Disposed database engines")

# Log Upload
async def upload_log_file(file_id: str, log_filename: str, logger: logging.Logger) -> Optional[str]:
    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        retry=retry_if_exception_type(Exception),
        before_sleep=lambda retry_state: logger.info(
            f"Retrying log upload for FileID {file_id} (attempt {retry_state.attempt_number}/3)"
        )
    )
    async def try_upload():
        if not os.path.exists(log_filename):
            logger.warning(f"Log file {log_filename} does not exist, skipping upload")
            return None

        file_hash = await asyncio.to_thread(hashlib.md5, open(log_filename, "rb").read())
        file_hash = file_hash.hexdigest()
        current_time = time.time()
        key = (log_filename, file_id)

        LAST_UPLOAD = {}
        if key in LAST_UPLOAD and LAST_UPLOAD[key]["hash"] == file_hash and current_time - LAST_UPLOAD[key]["time"] < 60:
            logger.info(f"Skipping redundant upload for {log_filename}")
            return LAST_UPLOAD[key]["url"]

        try:
            upload_url = await upload_file_to_space(
                file_src=log_filename,
                save_as=f"job_logs/job_{file_id}.log",
                is_public=True,
                logger=logger,
                file_id=file_id
            )
            if not upload_url:
                logger.error(f"S3 upload returned empty URL for {log_filename}")
                raise ValueError("Empty upload URL")
            await update_log_url_in_db(file_id, upload_url, logger)
            LAST_UPLOAD[key] = {"hash": file_hash, "time": current_time, "url": upload_url}
            logger.info(f"Log uploaded to R2: {upload_url}")
            return upload_url
        except Exception as e:
            logger.error(f"Failed to upload log for FileID {file_id}: {e}", exc_info=True)
            raise

    try:
        return await try_upload()
    except Exception as e:
        logger.error(f"Failed to upload log for FileID {file_id} after retries: {e}", exc_info=True)
        return None

# Main FastAPI Endpoint (Optimized for Async Smoothness)
async def generate_download_file(
    file_id: int,
    background_tasks: BackgroundTasks = None,
    logger: Optional[logging.Logger] = None,
    file_id_param: Optional[int] = None,
    preferred_image_method: str = "append"
) -> Dict[str, str]:
    logger, log_filename = setup_job_logger(job_id=str(file_id), log_dir="job_logs", console_output=True)
    logger.setLevel(logging.DEBUG)  # Enable debug logging for diagnostics
    temp_images_dir, temp_excel_dir = None, None
    successful_entries = 0
    failed_entries = 0
    entries = []

    try:
        file_id = int(file_id)
        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("SELECT FileName FROM utb_ImageScraperFiles WHERE ID = :file_id"),
                {"file_id": file_id}
            )
            row = result.fetchone()
            result.close()
            if not row:
                logger.error(f"No file found for ID {file_id}")
                return {"error": f"No file found for ID {file_id}", "log_filename": log_filename}
            original_filename = row[0]
        logger.debug(f"Original filename: {original_filename}")

        logger.info(f"Fetching images for ID: {file_id}")
        selected_images_df = await get_images_excel_db(str(file_id), logger=logger)
        logger.info(f"Fetched DataFrame for ID {file_id}, shape: {selected_images_df.shape}")
        logger.debug(f"DataFrame contents:\n{selected_images_df.to_string()}")

        expected_columns = ["ExcelRowID", "ImageUrl", "ImageUrlThumbnail", "Brand", "Style", "Color", "Category"]
        if list(selected_images_df.columns) != expected_columns:
            logger.error(f"Invalid columns in DataFrame for ID {file_id}. Got: {list(selected_images_df.columns)}, Expected: {expected_columns}")
            return {"error": f"Invalid DataFrame columns", "log_filename": log_filename}

        if selected_images_df.empty:
            logger.error(f"DataFrame is empty for FileID {file_id}")
            return {"error": f"No image data found for FileID {file_id}", "log_filename": log_filename}

        entries = selected_images_df['ExcelRowID'].unique().tolist()
        logger.info(f"Processing {len(entries)} unique entries for FileID {file_id}")

        template_file_path = "https://iconluxurygroup.s3.us-east-2.amazonaws.com/ICON_DISTRO_USD_20250312.xlsx"
        base_name, extension = os.path.splitext(original_filename)
        processed_file_name = f"excel_files/{file_id}/{base_name}{extension}"

        temp_images_dir, temp_excel_dir = await create_temp_dirs(file_id, logger=logger)
        logger.debug(f"Created temp directories: images={temp_images_dir}, excel={temp_excel_dir}")
        local_filename = os.path.join(temp_excel_dir, original_filename)

        async with httpx.AsyncClient(timeout=httpx.Timeout(30, connect=10)) as client:
            logger.debug(f"Downloading template from {template_file_path}")
            try:
                response = await client.get(template_file_path)
                response.raise_for_status()
                async with aiofiles.open(local_filename, 'wb') as f:
                    content = await response.read()
                    if not content:
                        logger.error(f"Empty content downloaded from {template_file_path}")
                        return {"error": f"Failed to download template file: empty content", "log_filename": log_filename}
                    await f.write(content)
            except httpx.HTTPStatusError as e:
                logger.error(f"HTTP error downloading template: {e}")
                return {"error": f"Failed to download template file: {e}", "log_filename": log_filename}
            except httpx.RequestError as e:
                logger.error(f"Request error downloading template: {e}")
                return {"error": f"Failed to download template file: {e}", "log_filename": log_filename}

        if not os.path.exists(local_filename):
            logger.error(f"Template file not found at {local_filename}")
            return {"error": f"Failed to download template file", "log_filename": log_filename}
        template_size = os.path.getsize(local_filename)
        logger.debug(f"Template file saved: {local_filename}, size: {template_size} bytes")
        if template_size < 1000:
            logger.warning(f"Template file {local_filename} is suspiciously small ({template_size} bytes)")

        has_valid_images = not selected_images_df.empty and any(pd.notna(row['ImageUrl']) and row['ImageUrl'] for _, row in selected_images_df.iterrows())
        
        if not has_valid_images:
            logger.warning(f"No valid images found for ID {file_id}")
            try:
                async with aiofiles.open(local_filename, 'rb') as f:
                    wb = await asyncio.to_thread(load_workbook, f)
                ws = wb.create_sheet("NoImages") if "NoImages" not in wb else wb["NoImages"]
                ws["A1"] = f"No valid images found for FileID {file_id}."
                await asyncio.to_thread(wb.save, local_filename)
            except Exception as e:
                logger.error(f"Failed to write no-images message to Excel: {e}", exc_info=True)
                return {"error": f"Failed to write Excel file: {e}", "log_filename": log_filename}
            failed_entries = len(entries)
        else:
            selected_image_list = [
                {
                    'ExcelRowID': int(row['ExcelRowID']),
                    'ImageUrl': row['ImageUrl'],
                    'ImageUrlThumbnail': row.get('ImageUrlThumbnail', ''),
                    'Brand': row.get('Brand', ''),
                    'Style': row.get('Style', ''),
                    'Color': row.get('Color', ''),
                    'Category': row.get('Category', ''),
                    'FileID': file_id
                }
                for _, row in selected_images_df.iterrows()
                if pd.notna(row['ImageUrl']) and row['ImageUrl']
            ]
            logger.info(f"Selected {len(selected_image_list)} valid images")
            if not selected_image_list:
                logger.error("No valid images after filtering, despite has_valid_images=True")
                return {"error": "No valid images after filtering", "log_filename": log_filename}

            try:
                success = await process_images_and_anchor(
                    selected_image_list,
                    temp_images_dir,
                    local_filename,
                    preferred_image_method=preferred_image_method,
                    logger=logger
                )
                if not success:
                    logger.error("Image processing pipeline reported failure")
            except Exception as e:
                logger.error(f"Image processing pipeline failed: {e}", exc_info=True)
                return {"error": f"Image processing failed: {e}", "log_filename": log_filename}
            
            successful_entries = len(entries) - sum(1 for item in selected_image_list if not os.path.exists(os.path.join(temp_images_dir, f"{item['ExcelRowID']}_image.jpg")))
            failed_entries = len(entries) - successful_entries
            logger.info(f"Processed {successful_entries} successful entries, {failed_entries} failed entries")

        if not os.path.exists(local_filename):
            logger.error(f"Excel file not found at {local_filename} after processing")
            return {"error": f"Excel file not found", "log_filename": log_filename}
        final_size = os.path.getsize(local_filename)
        logger.debug(f"Excel file exists: {local_filename}, size: {final_size} bytes")
        if final_size < 1000:
            logger.warning(f"Final Excel file {local_filename} is suspiciously small ({final_size} bytes)")

        logger.info(f"Uploading Excel file to R2: {processed_file_name}")
        try:
            public_url = await upload_file_to_space(
                file_src=local_filename,
                save_as=processed_file_name,
                is_public=True,
                logger=logger,
                file_id=file_id
            )
            if not public_url:
                logger.error(f"Upload failed for ID {file_id}: No public URL returned")
                return {"error": "Failed to upload processed file", "log_filename": log_filename}
            logger.info(f"Excel file uploaded to R2, public_url: {public_url}")
        except Exception as e:
            logger.error(f"S3 upload failed for {local_filename}: {e}", exc_info=True)
            return {"error": f"Failed to upload processed file: {e}", "log_filename": log_filename}

        try:
            await update_file_location_complete(str(file_id), public_url, logger=logger)
            await update_file_generate_complete(str(file_id), logger=logger)
        except Exception as e:
            logger.error(f"Failed to update database with file location: {e}", exc_info=True)

        log_public_url = await upload_log_file(str(file_id), log_filename, logger)
        if not log_public_url:
            logger.warning(f"Log upload failed for {log_filename}")

        send_to_email_addr = await get_send_to_email(file_id, logger=logger)
        if not send_to_email_addr:
            logger.error(f"No email address for ID {file_id}")
            return {"error": "Failed to retrieve email address", "log_filename": log_filename}

        subject_line = f"{original_filename} Job Notification{' - No Images' if not has_valid_images else ''}"
        user_message = (
            f"Excel file generation for FileID {file_id} {'completed successfully' if successful_entries > 0 else 'completed with no valid images'}.\n"
            f"Download Excel file: {public_url}\n"
            f"Log file: {log_public_url or log_filename}"
        )
        logger.debug(f"Scheduling user email to {send_to_email_addr} with public_url: {public_url}, message: {user_message}")
        if background_tasks:
            background_tasks.add_task(
                send_message_email,
                to_emails=send_to_email_addr,
                subject=subject_line,
                message=user_message,
                logger=logger
            )

        admin_email = "nik@luxurymarket.com"
        admin_subject = f"Admin Log: Job Completed for FileID {file_id}"
        admin_message = (
            f"Processing for FileID {file_id} completed.\n"
            f"Successful entries: {successful_entries}/{len(entries)}\n"
            f"Failed entries: {failed_entries}\n"
            f"Last EntryID: {max(entries) if entries else 0}\n"
            f"Excel file: {public_url}\n"
            f"Log file: {log_public_url or log_filename}"
        )
        logger.debug(f"Scheduling admin email to {admin_email} with message: {admin_message}")
        if background_tasks:
            background_tasks.add_task(
                send_message_email,
                to_emails=[admin_email],
                subject=admin_subject,
                message=admin_message,
                logger=logger
            )

        logger.info(f"Completed ID {file_id}")
        return {
            "message": "Processing completed successfully" if successful_entries > 0 else "No valid images found, empty Excel file generated",
            "public_url": public_url,
            "log_filename": log_filename,
            "log_public_url": log_public_url or ""
        }
    except Exception as e:
        logger.error(f"Error for ID {file_id}: {e}", exc_info=True)
        log_public_url = await upload_log_file(str(file_id), log_filename, logger)
        send_to_email_addr = await get_send_to_email(file_id, logger=logger)
        if send_to_email_addr and background_tasks:
            error_message = f"Excel file generation for FileID {file_id} failed.\nError: {str(e)}\nLog file: {log_public_url or log_filename}"
            logger.debug(f"Scheduling error email to {send_to_email_addr} with message: {error_message}")
            background_tasks.add_task(
                send_message_email,
                to_emails=send_to_email_addr,
                subject=f"Error: Job Failed for FileID {file_id}",
                message=error_message,
                logger=logger
            )
        return {"error": f"An error occurred: {str(e)}", "log_filename": log_filename, "log_public_url": log_public_url or ""}
    finally:
        if temp_images_dir and temp_excel_dir:
            try:
                await cleanup_temp_dirs([temp_images_dir, temp_excel_dir], logger=logger)
            except Exception as e:
                logger.error(f"Failed to clean up temp directories: {e}", exc_info=True)
        await async_engine.dispose()
        logger.info(f"Cleaned up resources for ID {file_id}")
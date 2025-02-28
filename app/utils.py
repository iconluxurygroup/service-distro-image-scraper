import json
import logging
import pyodbc
from config import conn_str
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

def clean_json(value):
    if not value or not isinstance(value, str) or value.strip() in ["None", "null", "NaN", "undefined"]:
        return json.dumps({
            "description": "",
            "user_provided": {"brand": "", "category": "", "color": ""},
            "extracted_features": {"brand": "", "category": "", "color": ""},
            "match_score": None,
            "reasoning_match": "",
            "linesheet_score": None,
            "reasoning_linesheet": ""
        })
    try:
        parsed = json.loads(value)
        return json.dumps(parsed)
    except json.JSONDecodeError:
        logging.error(f"Invalid JSON value: {value}")
        return json.dumps({"error": "Invalid JSON"})

def get_endpoint():
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            sql_query = "SELECT TOP 1 EndpointURL FROM utb_Endpoints WHERE EndpointIsBlocked = 0 ORDER BY NEWID()"
            cursor.execute(sql_query)
            endpoint_url = cursor.fetchone()
            if endpoint_url:
                return endpoint_url[0]
            return "No EndpointURL"
    except Exception as e:
        logging.error(f"Error getting endpoint URL: {e}")
        return "No EndpointURL"

def remove_endpoint(endpoint):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            sql_query = f"UPDATE utb_Endpoints SET EndpointIsBlocked = 1 WHERE EndpointURL = '{endpoint}'"
            cursor.execute(sql_query)
            conn.commit()
            logging.info(f"Marked endpoint as blocked: {endpoint}")
    except Exception as e:
        logging.error(f"Error marking endpoint as blocked: {e}")

def prepare_images_for_download_dataframe(df):
    images_to_download = []
    try:
        for row in df.itertuples(index=False, name=None):
            if row[1] != 'No google image results found':
                images_to_download.append(row)
        logging.info(f"Prepared {len(images_to_download)} images for download")
        return images_to_download
    except Exception as e:
        logging.error(f"Error preparing images for download: {e}")
        return []

def write_failed_downloads_to_excel(failed_downloads, excel_file):
    from openpyxl import load_workbook
    from excel_utils import highlight_cell
    if failed_downloads:
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            for row in failed_downloads:
                url, row_id = row
                if url and url != 'None found in this filter':
                    cell_reference = f"{get_column_letter(1)}{row_id}"
                    worksheet[cell_reference] = str(url)
                    highlight_cell(excel_file, cell_reference)
            workbook.save(excel_file)
            logging.info(f"Failed downloads written to Excel file: {excel_file}")
            return True
        except Exception as e:
            logging.error(f"Error writing failed downloads to Excel: {e}")
            return False
    else:
        logging.info("No failed downloads to write to Excel.")
        return True

def check_json_status(file_id):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            total_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ?
            """
            cursor.execute(total_query, (file_id,))
            total_count = cursor.fetchone()[0]
            
            if total_count == 0:
                return {"file_id": file_id, "status": "no_records", "message": "No records found for this file ID"}
            
            issues_data = {}
            null_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? AND t.aijson IS NULL
            """
            cursor.execute(null_query, (file_id,))
            issues_data["null_json"] = cursor.fetchone()[0]
            
            empty_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? AND t.aijson = ''
            """
            cursor.execute(empty_query, (file_id,))
            issues_data["empty_json"] = cursor.fetchone()[0]
            
            total_issues = sum(val for val in issues_data.values() if isinstance(val, int))
            issue_percentage = (total_issues / total_count * 100) if total_count > 0 else 0
            
            return {
                "file_id": file_id,
                "total_records": total_count,
                "total_issues": total_issues,
                "issue_percentage": round(issue_percentage, 2),
                "status": "needs_fixing" if issue_percentage > 0 else "healthy",
                "issue_breakdown": issues_data,
                "sample_issues": []
            }
    except Exception as e:
        logging.error(f"Error checking JSON status: {e}")
        return {"file_id": file_id, "status": "error", "error_message": str(e)}

def fix_json_data(background_tasks, file_id=None, limit=1000):
    try:
        def background_fix_json():
            with pyodbc.connect(conn_str) as conn:
                cursor = conn.cursor()
                if file_id:
                    query = f"""
                        SELECT TOP {limit} t.ResultID, t.aijson 
                        FROM utb_ImageScraperResult t
                        INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                        WHERE r.FileID = ? AND (t.aijson IS NULL OR ISJSON(t.aijson) = 0 OR t.aijson = '')
                    """
                    cursor.execute(query, (file_id,))
                else:
                    query = f"""
                        SELECT TOP {limit} t.ResultID, t.aijson 
                        FROM utb_ImageScraperResult t
                        WHERE t.aijson IS NULL OR ISJSON(t.aijson) = 0 OR t.aijson = ''
                    """
                    cursor.execute(query)
                
                rows = cursor.fetchall()
                total_fixed = 0
                for result_id, aijson_value in rows:
                    cleaned_json = clean_json(aijson_value)
                    cursor.execute("UPDATE utb_ImageScraperResult SET aijson = ? WHERE ResultID = ?", (cleaned_json, result_id))
                    total_fixed += 1
                conn.commit()
                logging.info(f"Fixed {total_fixed} records")
                return {"status": "completed", "records_fixed": total_fixed}
        
        background_tasks.add_task(background_fix_json)
        return {"message": f"JSON fix operation initiated in background" + (f" for FileID: {file_id}" if file_id else " across all files"), "status": "processing", "limit": limit}
    except Exception as e:
        logging.error(f"Error initiating JSON fix operation: {e}")
        return {"error": f"An error occurred: {str(e)}"}
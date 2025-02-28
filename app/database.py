import pyodbc
import pandas as pd
import logging
import time
import traceback,requests,urllib
import re, chardet
from icon_image_lib.LR import LR
from fastapi import FastAPI, BackgroundTasks
import asyncio, os, threading, uuid, requests, openpyxl, uvicorn, shutil, mimetypes, time
from openpyxl import load_workbook
from PIL import Image as IMG2
from PIL import UnidentifiedImageError
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import datetime, re
import boto3
import logging
from io import BytesIO
from openpyxl.utils import get_column_letter
from icon_image_lib.google_parser import get_original_images as GP
from requests.adapters import HTTPAdapter
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib3.util.retry import Retry
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition, Personalization, Cc, To
import base64, zlib
import json
import ray
import tldextract
from collections import Counter

import pandas as pd  # Ensure this is present for DataFrame
import pyodbc  # Ensure this is present for database connection
import math
import asyncio
from fastapi import BackgroundTasks
from image_processing import get_image_data, analyze_image_with_grok_vision
from config import conn_str,engine
logger = logging.getLogger(__name__)

def fetch_pending_images(limit=10):
    query = """
        SELECT rr.ResultID, rr.EntryID, rr.ImageURL, 
               r.ProductBrand, r.ProductCategory, r.ProductColor
        FROM utb_ImageScraperRecords r
        INNER JOIN utb_ImageScraperResult rr ON r.EntryID = rr.EntryID
        WHERE rr.aijson IS NULL
        ORDER BY rr.ResultID
    """
    try:
        with pyodbc.connect(conn_str) as conn:
            df = pd.read_sql(query, conn, params=[limit])
            logging.info(f"Fetched {len(df)} pending images for processing")
            return df
    except Exception as e:
        logging.error(f"Error fetching pending images: {e}")
        return pd.DataFrame()

def insert_file_db(file_name, file_source, send_to_email="nik@iconluxurygroup.com"):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            insert_query = """
                INSERT INTO utb_ImageScraperFiles (FileName, FileLocationUrl, UserEmail) 
                OUTPUT INSERTED.Id 
                VALUES (?, ?, ?)
            """
            values = (file_name, file_source, send_to_email)
            cursor.execute(insert_query, values)
            file_id = cursor.fetchval()
            conn.commit()
            logging.info(f"Inserted new file record with ID: {file_id}")
            return file_id
    except Exception as e:
        logging.error(f"Error inserting file record: {e}")
        raise

def update_database(result_id, ai_json, ai_caption=None):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            update_query = """
                UPDATE utb_ImageScraperResult 
                SET aijson = ?, aicaption = ? 
                WHERE ResultID = ?
            """
            cursor.execute(update_query, (ai_json, ai_caption, result_id))
            conn.commit()
            logging.info(f"Updated database for ResultID: {result_id}")
    except Exception as e:
        logging.error(f"Error updating database: {e}")
        raise
def load_payload_db(rows, file_id):
    try:
        file_id = int(file_id)
        with pyodbc.connect(conn_str) as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
            if cursor.fetchone()[0] == 0:
                raise ValueError(f"FileID {file_id} does not exist in utb_ImageScraperFiles")
            
            df = pd.DataFrame(rows).rename(columns={
                'absoluteRowIndex': 'ExcelRowID',
                'searchValue': 'ProductModel',
                'brandValue': 'ProductBrand',
                'colorValue': 'ProductColor',
                'CategoryValue': 'ProductCategory'
            })
            
            # Validate that ProductBrand is present
            if 'ProductBrand' not in df.columns or df['ProductBrand'].isnull().all():
                logging.warning(f"No valid ProductBrand data found in payload for FileID {file_id}")
                df['ProductBrand'] = df.get('ProductBrand', '')  # Default to empty string if missing
            
            df.insert(0, 'FileID', file_id)
            if 'imageValue' in df.columns:
                df.drop(columns=['imageValue'], inplace=True)
            
            logging.debug(f"Inserting data for FileID {file_id}: {df[['ProductBrand', 'ProductModel']].head().to_dict()}")
            for _, row in df.iterrows():
                cursor.execute(
                    f"INSERT INTO utb_ImageScraperRecords ({', '.join(df.columns)}) VALUES ({', '.join(['?'] * len(df.columns))})",
                    tuple(row)
                )
            connection.commit()
        logging.info(f"Loaded {len(df)} rows into utb_ImageScraperRecords for FileID: {file_id}")
        return df
    except Exception as e:
        logging.error(f"Error loading payload data: {e}")
        raise
def unpack_content(encoded_content):
    """
    Unpack base64 encoded and compressed content.
    
    Args:
        encoded_content (str): Base64 encoded and compressed content
        
    Returns:
        bytes: Unpacked content
    """
    try:
        if encoded_content:
            compressed_content = base64.b64decode(encoded_content)
            original_content = zlib.decompress(compressed_content)
            return original_content  # Return as binary data
        return None
    except Exception as e:
        logging.error(f"Error unpacking content: {e}")
        return None

def get_records_to_search(file_id):
    """Return two search strings per EntryID: one with ProductModel, one with ProductModel + ProductBrand."""
    try:
        file_id = int(file_id)
        sql_query = """
            SELECT EntryID, ProductModel AS SearchString, 'model_only' AS SearchType, FileID
            FROM utb_ImageScraperRecords 
            WHERE FileID = ? AND Step1 IS NULL
            UNION ALL
            SELECT EntryID, 
                   CASE 
                       WHEN ProductBrand IS NULL OR ProductBrand = '' 
                       THEN ProductModel 
                       ELSE ProductModel + ' ' + ProductBrand 
                   END AS SearchString, 
                   'model_plus_brand' AS SearchType, 
                   FileID
            FROM utb_ImageScraperRecords 
            WHERE FileID = ? AND Step1 IS NULL
            ORDER BY EntryID, SearchType
        """
        with pyodbc.connect(conn_str) as connection:
            df = pd.read_sql_query(sql_query, connection, params=[file_id, file_id])
        
        # Validate FileID consistency
        if not df.empty:
            invalid_rows = df[df['FileID'] != file_id]
            if not invalid_rows.empty:
                logger.error(f"Found {len(invalid_rows)} rows with incorrect FileID for requested FileID {file_id}: {invalid_rows[['EntryID', 'FileID']].to_dict()}")
                df = df[df['FileID'] == file_id]
        
        logger.info(f"Got {len(df)} search records (2 per EntryID) for FileID: {file_id}")
        return df[['EntryID', 'SearchString', 'SearchType']]
    except Exception as e:
        logger.error(f"Error getting records to search for FileID {file_id}: {e}")
        return pd.DataFrame()
    
def process_search_row(search_string, endpoint, entry_id):
    try:
        # Ensure search string is valid and includes brand context
        if not search_string or len(search_string.strip()) < 3:
            logging.warning(f"Invalid search string for EntryID {entry_id}: '{search_string}'")
            return False
        
        search_url = f"{endpoint}?query={urllib.parse.quote(search_string + ' brand')}"  # Append 'brand' to prioritize
        logging.info(f"Searching URL: {search_url}")
        
        response = requests.get(search_url, timeout=60)
        if response.status_code != 200:
            logging.warning(f"Non-200 status {response.status_code} for {search_url}")
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            return process_search_row(search_string, n_endpoint, entry_id)
        
        try:
            response_json = response.json()
            result = response_json.get('body', None)
        except json.JSONDecodeError:
            logging.warning(f"Invalid JSON response for {search_url}")
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            return process_search_row(search_string, n_endpoint, entry_id)
        
        if not result:
            logging.warning(f"No body in response for {search_url}")
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            return process_search_row(search_string, n_endpoint, entry_id)
        
        unpacked_html = unpack_content(result)
        if not unpacked_html or len(unpacked_html) < 100:
            logging.warning(f"Invalid unpacked HTML for {search_url}")
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            return process_search_row(search_string, n_endpoint, entry_id)
        
        parsed_data = GP(unpacked_html)
        if not parsed_data or not isinstance(parsed_data, tuple) or not parsed_data[0]:
            logging.warning(f"No valid parsed data for {search_url}")
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            return process_search_row(search_string, n_endpoint, entry_id)
        
        urls, descriptions, sources, thumbnails = parsed_data
        if urls and urls[0] != 'No google image results found':
            df = pd.DataFrame({
                'EntryId': [entry_id] * len(urls),
                'ImageUrl': urls,
                'ImageDesc': descriptions,
                'ImageSource': sources,
                'ImageUrlThumbnail': thumbnails
            })
            df.to_sql(name='utb_ImageScraperResult', con=engine, index=False, if_exists='append')
            
            with pyodbc.connect(conn_str) as connection:
                cursor = connection.cursor()
                cursor.execute(f"UPDATE utb_ImageScraperRecords SET Step1 = GETDATE() WHERE EntryID = {entry_id}")
                connection.commit()
            logging.info(f"Processed EntryID {entry_id} with {len(df)} images")
            return True
        
        # If no valid image URL, retry
        logging.warning('No valid image URL, trying again with new endpoint')
        remove_endpoint(endpoint)
        n_endpoint = get_endpoint()
        return process_search_row(search_string, n_endpoint, entry_id)
    
    except requests.RequestException as e:
        logging.error(f"Request error: {e}")
        remove_endpoint(endpoint)
        n_endpoint = get_endpoint()
        logging.info(f"Trying again with new endpoint: {n_endpoint}")
        return process_search_row(search_string, n_endpoint, entry_id)
    
    except Exception as e:
        logging.error(f"Error processing search row: {e}")
        remove_endpoint(endpoint)
        n_endpoint = get_endpoint()
        logging.info(f"Trying again with new endpoint: {n_endpoint}")
        return process_search_row(search_string, n_endpoint, entry_id)
async def batch_process_images(file_id, limit):
    try:
        missing_df = fetch_missing_images(file_id=file_id, limit=limit, ai_analysis_only=True)
        if missing_df.empty:
            logging.info(f"No images need AI processing for FileID: {file_id}")
            return
        
        processed_count = 0
        for _, row in missing_df.iterrows():
            result_id = row['ResultID']
            image_url = row['ImageURL']
            try:
                image_data = get_image_data(image_url)
                if image_data:
                    features = await analyze_image_with_grok_vision(image_data)
                    if features and "error" not in features:  # Check for valid response
                        ai_json = json.dumps(features)
                    else:
                        logger.warning(f"Invalid Grok Vision response for ResultID {result_id}: {features}")
                        ai_json = json.dumps({
                            "description": "Failed to analyze",
                            "user_provided": {"brand": row['ProductBrand'], "category": row['ProductCategory'], "color": row['ProductColor']},
                            "extracted_features": {"brand": "", "category": "", "color": ""},
                            "match_score": None,
                            "reasoning_match": "API failure",
                            "linesheet_score": None,
                            "reasoning_linesheet": "API failure"
                        })
                    update_database(result_id, ai_json, "AI analysis failed")
                    processed_count += 1
                else:
                    logger.warning(f"No image data retrieved for ResultID {result_id}")
            except Exception as e:
                logger.error(f"Failed to process image for ResultID {result_id} (URL: {image_url}): {e}")
                ai_json = json.dumps({
                    "description": "Download or API error",
                    "user_provided": {"brand": row['ProductBrand'], "category": row['ProductCategory'], "color": row['ProductColor']},
                    "extracted_features": {"brand": "", "category": "", "color": ""},
                    "match_score": None,
                    "reasoning_match": str(e),
                    "linesheet_score": None,
                    "reasoning_linesheet": str(e)
                })
                update_database(result_id, ai_json, "AI analysis failed")
                continue
        
        logging.info(f"Processed {processed_count} out of {len(missing_df)} images for AI analysis for FileID: {file_id}")
    except Exception as e:
        logging.error(f"Error in batch_process_images: {e}")
        raise

async def process_images(file_id):
    try:
        images_df = fetch_images_by_file_id(file_id)
        if images_df.empty:
            logging.info(f"No images found for FileID: {file_id}")
            return
        
        processed_count = 0
        for _, row in images_df.iterrows():
            result_id = row['ResultID']
            image_url = row['ImageURL']
            if not row.get('aijson'):
                try:
                    image_data = get_image_data(image_url)
                    if image_data:
                        features = await analyze_image_with_grok_vision(image_data)
                        ai_json = json.dumps(features)
                        update_database(result_id, ai_json, "AI-generated caption")
                        processed_count += 1
                except Exception as e:
                    logging.error(f"Failed to process image for ResultID {result_id} (URL: {image_url}): {e}")
                    continue
        
        logging.info(f"Completed image processing for FileID: {file_id}, processed {processed_count} images")
    except Exception as e:
        logging.error(f"Error in process_images: {e}")
        raise

def fetch_images_by_file_id(file_id):
    query = """
        SELECT rr.ResultID, rr.EntryID, rr.ImageURL, 
               r.ProductBrand, r.ProductCategory, r.ProductColor,
               rr.aijson
        FROM utb_ImageScraperRecords r
        INNER JOIN utb_ImageScraperResult rr ON r.EntryID = rr.EntryID
        WHERE r.FileID = ?
    """
    try:
        with pyodbc.connect(conn_str) as conn:
            df = pd.read_sql(query, conn, params=[file_id])
            logging.info(f"Fetched {len(df)} images for FileID {file_id}")
            return df
    except Exception as e:
        logging.error(f"Error fetching images for FileID {file_id}: {e}")
        return pd.DataFrame()
def fetch_missing_images(file_id=None, limit=8, ai_analysis_only=True):
    """
    Fetch images with missing or NaN JSON fields from the database.
    
    Args:
        file_id (int, optional): The FileID to fetch missing images for
        limit (int, optional): Maximum number of records to fetch
        ai_analysis_only (bool, optional): If True, only fetch images missing AI analysis but with URLs.
                                          If False, fetch all missing records including those without URLs.
        
    Returns:
        pd.DataFrame: DataFrame containing missing image records
    """
    try:
        connection = pyodbc.connect(conn_str)
        query_params = []

        # Base SQL query depends on whether we're looking for just AI analysis issues or also missing URLs
        if ai_analysis_only:
            # Only looking for images with URLs but missing AI analysis
            query = """
            SELECT t.ResultID, t.EntryID, t.ImageURL, r.ProductBrand, r.ProductCategory, r.ProductColor
            FROM utb_ImageScraperResult t
            INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
            WHERE (
                t.ImageURL IS NOT NULL 
                AND t.ImageURL <> ''
                AND (
                    ISJSON(t.aijson) = 0 
                    OR t.aijson IS NULL
                    OR JSON_VALUE(t.aijson, '$.linesheet_score') IS NULL
                    OR JSON_VALUE(t.aijson, '$.linesheet_score') IN ('NaN', 'null', 'undefined')
                    OR JSON_VALUE(t.aijson, '$.match_score') IS NULL
                    OR JSON_VALUE(t.aijson, '$.match_score') IN ('NaN', 'null', 'undefined')
                )
            )
            """
        else:
            # Looking for records missing either URLs or AI analysis
            query = """
            -- Records missing in result table completely
            SELECT NULL as ResultID, r.EntryID, NULL as ImageURL, 
                   r.ProductBrand, r.ProductCategory, r.ProductColor
            FROM utb_ImageScraperRecords r
            LEFT JOIN utb_ImageScraperResult t ON r.EntryID = t.EntryID
            WHERE t.EntryID IS NULL
            
            UNION ALL
            
            -- Records with empty or NULL ImageURL
            SELECT t.ResultID, t.EntryID, t.ImageURL, 
                   r.ProductBrand, r.ProductCategory, r.ProductColor
            FROM utb_ImageScraperResult t
            INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
            WHERE t.ImageURL IS NULL OR t.ImageURL = ''
            
            UNION ALL
            
            -- Records with URLs but missing AI analysis
            SELECT t.ResultID, t.EntryID, t.ImageURL, 
                   r.ProductBrand, r.ProductCategory, r.ProductColor
            FROM utb_ImageScraperResult t
            INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
            WHERE t.ImageURL IS NOT NULL 
              AND t.ImageURL <> ''
              AND (
                  ISJSON(t.aijson) = 0 
                  OR t.aijson IS NULL
                  OR JSON_VALUE(t.aijson, '$.linesheet_score') IS NULL
                  OR JSON_VALUE(t.aijson, '$.linesheet_score') IN ('NaN', 'null', 'undefined')
                  OR JSON_VALUE(t.aijson, '$.match_score') IS NULL
                  OR JSON_VALUE(t.aijson, '$.match_score') IN ('NaN', 'null', 'undefined')
              )
            """

        # Add FileID filter if provided
        if file_id:
            query += " AND r.FileID = ?"
            query_params.append(file_id)

        # Limit number of results
        query += " ORDER BY r.EntryID ASC OFFSET 0 ROWS FETCH NEXT ? ROWS ONLY"
        query_params.append(limit)

        # Fetch data
        df = pd.read_sql_query(query, connection, params=query_params)
        connection.close()
        
        if df.empty:
            logger.info(f"No missing images found" + (f" for FileID: {file_id}" if file_id else ""))
        else:
            logger.info(f"Found {len(df)} missing images" + (f" for FileID: {file_id}" if file_id else ""))
            
        return df

    except Exception as e:
        logger.error(f"Error fetching missing images: {e}")
        return pd.DataFrame()
def update_initial_sort_order(file_id):
    """Set initial SortOrder based on ResultID after fetching image results."""
    try:
        file_id = int(file_id)
        with pyodbc.connect(conn_str) as connection:
            connection.timeout = 300
            cursor = connection.cursor()
            logger.info(f"🔄 Setting initial SortOrder for FileID: {file_id}")
            
            # Begin transaction
            cursor.execute("BEGIN TRANSACTION")
            
            # Reset Step1 in utb_ImageScraperRecords to allow reprocessing
            cursor.execute("UPDATE utb_ImageScraperRecords SET Step1 = NULL WHERE FileID = ?", (file_id,))
            reset_step1_count = cursor.rowcount
            logger.info(f"Reset Step1 for {reset_step1_count} rows in utb_ImageScraperRecords")
            
            # Reset existing SortOrder in utb_ImageScraperResult
            cursor.execute("UPDATE utb_ImageScraperResult SET SortOrder = NULL WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = ?)", (file_id,))
            reset_sort_count = cursor.rowcount
            logger.info(f"Reset SortOrder for {reset_sort_count} rows in utb_ImageScraperResult")
            
            # Set initial SortOrder based on ResultID
            initial_sort_query = """
                WITH toupdate AS (
                    SELECT t.*, 
                           ROW_NUMBER() OVER (PARTITION BY t.EntryID ORDER BY t.ResultID) AS seqnum
                    FROM utb_ImageScraperResult t 
                    INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID 
                    WHERE r.FileID = ?
                ) 
                UPDATE toupdate 
                SET SortOrder = seqnum;
            """
            cursor.execute(initial_sort_query, (file_id,))
            update_count = cursor.rowcount
            logger.info(f"Set initial SortOrder for {update_count} rows")
            
            # Commit transaction
            cursor.execute("COMMIT")
            
            # Verify results
            verify_query = """
                SELECT t.ResultID, t.EntryID, t.SortOrder
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ?
                ORDER BY t.EntryID, t.SortOrder
            """
            cursor.execute(verify_query, (file_id,))
            results = cursor.fetchall()
            for record in results:
                logger.info(f"Initial - EntryID: {record[1]}, ResultID: {record[0]}, SortOrder: {record[2]}")
            
            # Check total result count
            count_query = "SELECT COUNT(*) FROM utb_ImageScraperResult WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = ?)"
            cursor.execute(count_query, (file_id,))
            total_results = cursor.fetchone()[0]
            logger.info(f"Total results after initial sort for FileID {file_id}: {total_results}")
            if total_results < 16:  # Assuming 1 EntryID
                logger.warning(f"Expected at least 16 results (8 per search type), got {total_results}")
            
            return [{"ResultID": row[0], "EntryID": row[1], "SortOrder": row[2]} for row in results]
    except Exception as e:
        logger.error(f"Error setting initial SortOrder: {e}")
        if 'cursor' in locals():
            cursor.execute("ROLLBACK")
        return None
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()  
def update_search_sort_order(file_id):   
    """Set initial SortOrder based on ResultID after fetching image results."""
    try:
        file_id = int(file_id)
        with pyodbc.connect(conn_str) as connection:  
            connection.timeout = 300
            cursor = connection.cursor()
            logger.info(f"🔄 Setting initial SortOrder for FileID: {file_id}")
                   
            # Set initial SortOrder based on ResultID
            initial_sort_query = """
                WITH toupdate AS (
                    SELECT t.*, 
                        ROW_NUMBER() OVER (PARTITION BY t.EntryID ORDER BY t.ResultID) AS seqnum
                    FROM utb_ImageScraperResult t 
                    INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID 
                    WHERE r.FileID = ?
                ) 
                UPDATE toupdate 
                SET SortOrder = seqnum;
            """
            cursor.execute("BEGIN TRANSACTION")
            cursor.execute(initial_sort_query, (file_id,))
            update_count = cursor.rowcount
            cursor.execute("COMMIT")
            logger.info(f"Set Search SortOrder for {update_count} rows")
            
            # Verify
            verify_query = """
                SELECT TOP 20 t.ResultID, t.EntryID, t.SortOrder
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ?
                ORDER BY t.EntryID, t.SortOrder
            """
            cursor.execute(verify_query, (file_id,))
            results = cursor.fetchall()
            for record in results:
                logger.info(f"Search - EntryID: {record[1]}, ResultID: {record[0]}, SortOrder: {record[2]}")
            
            return [{"ResultID": row[0], "EntryID": row[1], "SortOrder": row[2]} for row in results]
    except Exception as e:
        logger.error(f"Error setting Search SortOrder: {e}")
        return None
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

def update_ai_sort_order(file_id):
    """Update SortOrder based on AI analysis scores."""
    try:
        file_id = int(file_id)
        with pyodbc.connect(conn_str) as connection:
            connection.timeout = 300
            cursor = connection.cursor()
            logger.info(f"🔄 Updating AI-based SortOrder for FileID: {file_id}")
            
            # Debug current state
            debug_query = """
                SELECT TOP 20 t.ResultID, t.EntryID, t.SortOrder, t.aijson,
                       CASE WHEN ISJSON(t.aijson) = 1 THEN JSON_VALUE(t.aijson, '$.match_score') ELSE '-1' END AS match_score,
                       CASE WHEN ISJSON(t.aijson) = 1 THEN JSON_VALUE(t.aijson, '$.linesheet_score') ELSE '-1' END AS linesheet_score
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ?
                ORDER BY t.EntryID, t.SortOrder
            """
            cursor.execute(debug_query, (file_id,))
            current_order = cursor.fetchall()
            for record in current_order:
                logger.info(f"Pre-AI - EntryID: {record[1]}, ResultID: {record[0]}, SortOrder: {record[2]}, MatchScore: {record[4]}, LinesheetScore: {record[5]}, aijson: {record[3][:100] if record[3] else 'None'}")
            
            # Update SortOrder based on AI scores
            ai_sort_query = """
                WITH RankedResults AS (
                    SELECT 
                        t.ResultID,
                        t.EntryID,
                        ROW_NUMBER() OVER (
                            PARTITION BY t.EntryID 
                            ORDER BY 
                                CASE WHEN ISJSON(t.aijson) = 1 AND ISNUMERIC(JSON_VALUE(t.aijson, '$.match_score')) = 1 
                                     THEN CAST(JSON_VALUE(t.aijson, '$.match_score') AS FLOAT) ELSE -1 END DESC,
                                CASE WHEN ISJSON(t.aijson) = 1 AND ISNUMERIC(JSON_VALUE(t.aijson, '$.linesheet_score')) = 1 
                                     THEN CAST(JSON_VALUE(t.aijson, '$.linesheet_score') AS FLOAT) ELSE -1 END DESC,
                                t.ResultID ASC
                        ) AS rank
                    FROM utb_ImageScraperResult t
                    INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                    WHERE r.FileID = ?
                )
                UPDATE utb_ImageScraperResult
                SET SortOrder = rr.rank
                FROM utb_ImageScraperResult t
                INNER JOIN RankedResults rr ON t.ResultID = rr.ResultID;
            """
            cursor.execute("BEGIN TRANSACTION")
            cursor.execute(ai_sort_query, (file_id,))
            update_count = cursor.rowcount
            cursor.execute("COMMIT")
            logger.info(f"Updated AI SortOrder for {update_count} rows")
            
            # Verify
            cursor.execute(debug_query, (file_id,))
            updated_order = cursor.fetchall()
            for record in updated_order:
                logger.info(f"Post-AI - EntryID: {record[1]}, ResultID: {record[0]}, SortOrder: {record[2]}, MatchScore: {record[4]}, LinesheetScore: {record[5]}, aijson: {record[3][:100] if record[3] else 'None'}")
            
            return [{"ResultID": row[0], "EntryID": row[1], "SortOrder": row[2]} for row in updated_order]
    except Exception as e:
        logger.error(f"Error updating AI SortOrder: {e}")
        return None
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

def update_file_location_complete(file_id, file_location):
    file_id = int(file_id)
    try:
        with pyodbc.connect(conn_str) as connection:
            cursor = connection.cursor()
            cursor.execute("BEGIN TRANSACTION")
            cursor.execute("UPDATE utb_ImageScraperFiles SET FileLocationURLComplete = ? WHERE ID = ?", (file_location, file_id))
            cursor.execute("COMMIT")
            logging.info(f"Updated file location URL for FileID: {file_id}")
    except Exception as e:
        logging.error(f"Error updating file location URL: {e}")
        cursor.execute("ROLLBACK")
        raise

def update_file_generate_complete(file_id):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            query = "UPDATE utb_ImageScraperFiles SET CreateFileCompleteTime = GETDATE() WHERE ID = ?"
            cursor.execute(query, (file_id,))
            conn.commit()
            logging.info(f"Marked file generation as complete for FileID: {file_id}")
    except Exception as e:
        logging.error(f"Error updating file generation completion time: {e}")
        raise

def get_file_location(file_id):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            query = "SELECT FileLocationUrl FROM utb_ImageScraperFiles WHERE ID = ?"
            cursor.execute(query, (file_id,))
            file_location_url = cursor.fetchone()
            if file_location_url:
                logging.info(f"Got file location URL for FileID: {file_id}: {file_location_url[0]}")
                return file_location_url[0]
            logging.warning(f"No file location URL found for FileID: {file_id}")
            return "No File Found"
    except Exception as e:
        logging.error(f"Error getting file location URL: {e}")
        return "Error retrieving file location"

def get_send_to_email(file_id):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            query = "SELECT UserEmail FROM utb_ImageScraperFiles WHERE ID = ?"
            cursor.execute(query, (file_id,))
            send_to_email = cursor.fetchone()
            if send_to_email:
                logging.info(f"Got email address for FileID: {file_id}: {send_to_email[0]}")
                return send_to_email[0]
            logging.warning(f"No email address found for FileID: {file_id}")
            return "No Email Found"
    except Exception as e:
        logging.error(f"Error getting email address: {e}")
        return "nik@iconluxurygroup.com"

def get_images_excel_db(file_id):
    try:
        file_id = int(file_id)
        with pyodbc.connect(conn_str) as connection:
            cursor = connection.cursor()
            cursor.execute("UPDATE utb_ImageScraperFiles SET CreateFileStartTime = GETDATE() WHERE ID = ?", (file_id,))
            connection.commit()
            
            query = """
                SELECT s.ExcelRowID, r.ImageUrl, r.ImageUrlThumbnail 
                FROM utb_ImageScraperFiles f
                INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
                INNER JOIN utb_ImageScraperResult r ON r.EntryID = s.EntryID 
                WHERE f.ID = ? AND r.SortOrder = 1
                ORDER BY s.ExcelRowID
            """
            df = pd.read_sql_query(query, connection, params=[file_id])
            logging.info(f"Queried FileID: {file_id}, retrieved {len(df)} images for Excel export")
        return df
    except Exception as e:
        logging.error(f"Error getting images for Excel export: {e}")
        return pd.DataFrame()

def get_lm_products(file_id):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            query = f"EXEC usp_ImageScrapergetMatchFromRetail {file_id}"
            cursor.execute(query)
            conn.commit()
            logging.info(f"Executed stored procedure to match products for FileID: {file_id}")
    except Exception as e:
        logging.error(f"Error executing stored procedure to match products: {e}")

def get_endpoint():
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            sql_query = "SELECT TOP 1 EndpointURL FROM utb_Endpoints WHERE EndpointIsBlocked = 0 ORDER BY NEWID()"
            cursor.execute(sql_query)
            endpoint_url = cursor.fetchone()
            if endpoint_url:
                endpoint = endpoint_url[0]
                logging.info(f"Got endpoint URL: {endpoint}")
                return endpoint
            else:
                logging.warning("No endpoint URL found")
                return "No EndpointURL"
    except Exception as e:
        logging.error(f"Error getting endpoint URL: {e}")
        return "No EndpointURL"

def remove_endpoint(endpoint):
    try:
        with pyodbc.connect(conn_str) as conn:
            cursor = conn.cursor()
            sql_query = f"UPDATE utb_Endpoints SET EndpointIsBlocked = 1 WHERE EndpointURL = '{endpoint}'"
            cursor.execute(sql_query)
            conn.commit()
            logging.info(f"Marked endpoint as blocked: {endpoint}")
    except Exception as e:
        logging.error(f"Error marking endpoint as blocked: {e}")

def check_json_status(file_id):
    try:
        connection = pyodbc.connect(conn_str)
        cursor = connection.cursor()
        
        total_query = """
            SELECT COUNT(*) 
            FROM utb_ImageScraperResult t
            INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
            WHERE r.FileID = ?
        """
        cursor.execute(total_query, (file_id,))
        total_count = cursor.fetchone()[0]
        
        if total_count == 0:
            cursor.close()
            connection.close()
            return {
                "file_id": file_id,
                "status": "no_records",
                "message": "No records found for this file ID"
            }
        
        issues_data = {}
        
        try:
            null_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? AND t.aijson IS NULL
            """
            cursor.execute(null_query, (file_id,))
            issues_data["null_json"] = cursor.fetchone()[0]
        except Exception as e:
            logging.warning(f"Error in null_json query: {e}")
            issues_data["null_json"] = "error"
        
        try:
            empty_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? AND t.aijson = ''
            """
            cursor.execute(empty_query, (file_id,))
            issues_data["empty_json"] = cursor.fetchone()[0]
        except Exception as e:
            logging.warning(f"Error in empty_json query: {e}")
            issues_data["empty_json"] = "error"
        
        try:
            invalid_format_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? AND t.aijson IS NOT NULL AND ISJSON(t.aijson) = 0
            """
            cursor.execute(invalid_format_query, (file_id,))
            issues_data["invalid_format"] = cursor.fetchone()[0]
        except Exception as e:
            logging.warning(f"Error in invalid_format query: {e}")
            issues_data["invalid_format"] = "error"
        
        try:
            invalid_match_score_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? 
                AND t.aijson IS NOT NULL 
                AND ISJSON(t.aijson) = 1
                AND (
                    JSON_VALUE(t.aijson, '$.match_score') IS NULL
                    OR JSON_VALUE(t.aijson, '$.match_score') IN ('NaN', 'null', 'undefined')
                    OR ISNUMERIC(JSON_VALUE(t.aijson, '$.match_score')) = 0
                )
            """
            cursor.execute(invalid_match_score_query, (file_id,))
            issues_data["invalid_match_score"] = cursor.fetchone()[0]
        except Exception as e:
            logging.warning(f"Error in invalid_match_score query: {e}")
            issues_data["invalid_match_score"] = "error"
        
        try:
            invalid_linesheet_score_query = """
                SELECT COUNT(*) 
                FROM utb_ImageScraperResult t
                INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                WHERE r.FileID = ? 
                AND t.aijson IS NOT NULL 
                AND ISJSON(t.aijson) = 1
                AND (
                    JSON_VALUE(t.aijson, '$.linesheet_score') IS NULL
                    OR JSON_VALUE(t.aijson, '$.linesheet_score') IN ('NaN', 'null', 'undefined')
                    OR ISNUMERIC(JSON_VALUE(t.aijson, '$.linesheet_score')) = 0
                )
            """
            cursor.execute(invalid_linesheet_score_query, (file_id,))
            issues_data["invalid_linesheet_score"] = cursor.fetchone()[0]
        except Exception as e:
            logging.warning(f"Error in invalid_linesheet_score query: {e}")
            issues_data["invalid_linesheet_score"] = "error"
        
        total_issues = sum(val for val in issues_data.values() if isinstance(val, int))
        
        sample_query = """
            SELECT TOP 5 t.ResultID, t.aijson
            FROM utb_ImageScraperResult t
            INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
            WHERE r.FileID = ? AND (
                t.aijson IS NULL
                OR t.aijson = ''
                OR ISJSON(t.aijson) = 0
            )
        """
        try:
            cursor.execute(sample_query, (file_id,))
            samples = cursor.fetchall()
            sample_data = [{"ResultID": row[0], "aijson_prefix": str(row[1])[:100] if row[1] else None} for row in samples]
        except Exception as e:
            logging.warning(f"Error getting sample data: {e}")
            sample_data = [{"error": str(e)}]
        
        cursor.close()
        connection.close()
        
        issue_percentage = (total_issues / total_count * 100) if total_count > 0 else 0
        
        return {
            "file_id": file_id,
            "total_records": total_count,
            "total_issues": total_issues,
            "issue_percentage": round(issue_percentage, 2),
            "status": "needs_fixing" if issue_percentage > 0 else "healthy",
            "issue_breakdown": issues_data,
            "sample_issues": sample_data
        }
    except Exception as e:
        logging.error(f"Error checking JSON status: {e}")
        return {
            "file_id": file_id,
            "status": "error",
            "error_message": str(e)
        }

def fix_json_data(background_tasks: BackgroundTasks, file_id=None, limit=1000):
    def background_fix_json():
        try:
            connection = pyodbc.connect(conn_str)
            cursor = connection.cursor()
            
            if file_id:
                query = f"""
                    SELECT TOP {limit} t.ResultID, t.aijson 
                    FROM utb_ImageScraperResult t
                    INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                    WHERE r.FileID = ? AND (
                        t.aijson IS NULL
                        OR ISJSON(t.aijson) = 0 
                        OR t.aijson = ''
                        OR LEFT(t.aijson, 1) = '.'
                        OR LEFT(t.aijson, 1) = ','
                        OR JSON_VALUE(t.aijson, '$.match_score') IN ('NaN', 'null', 'undefined')
                        OR JSON_VALUE(t.aijson, '$.linesheet_score') IN ('NaN', 'null', 'undefined')
                    )
                """
                try:
                    cursor.execute(query, (file_id,))
                except Exception as e:
                    logging.warning(f"Error in complex query: {e}, falling back to simpler query")
                    query = f"""
                        SELECT TOP {limit} t.ResultID, t.aijson 
                        FROM utb_ImageScraperResult t
                        INNER JOIN utb_ImageScraperRecords r ON r.EntryID = t.EntryID
                        WHERE r.FileID = ?
                    """
                    cursor.execute(query, (file_id,))
            else:
                query = """
                    SELECT TOP ? t.ResultID, t.aijson 
                    FROM utb_ImageScraperResult t
                    WHERE 
                        t.aijson IS NULL
                        OR ISJSON(t.aijson) = 0 
                        OR t.aijson = ''
                        OR LEFT(t.aijson, 1) = '.'
                        OR LEFT(t.aijson, 1) = ','
                """
                try:
                    cursor.execute(query, (limit,))
                except Exception as e:
                    logging.warning(f"Error in complex query: {e}, falling back to simpler query")
                    query = "SELECT TOP ? ResultID, aijson FROM utb_ImageScraperResult"
                    cursor.execute(query, (limit,))
            
            rows = cursor.fetchall()
            logging.info(f"Found {len(rows)} records with potentially invalid JSON")
            
            batch_size = 100
            total_fixed = 0
            error_count = 0
            
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                updates = []
                
                for row in batch:
                    result_id, aijson_value = row
                    try:
                        cleaned_json = clean_json(aijson_value)
                        if cleaned_json:
                            updates.append((cleaned_json, result_id))
                    except Exception as e:
                        logging.error(f"Error cleaning JSON for ResultID {result_id}: {e}")
                        error_count += 1
                
                if updates:
                    try:
                        cursor.executemany(
                            "UPDATE utb_ImageScraperResult SET aijson = ? WHERE ResultID = ?",
                            updates
                        )
                        connection.commit()
                        batch_fixed = len(updates)
                        total_fixed += batch_fixed
                        logging.info(f"Fixed {batch_fixed} records in batch (total: {total_fixed})")
                    except Exception as batch_error:
                        logging.error(f"Error in batch update: {batch_error}")
                        for cleaned_json, result_id in updates:
                            try:
                                cursor.execute(
                                    "UPDATE utb_ImageScraperResult SET aijson = ? WHERE ResultID = ?",
                                    (cleaned_json, result_id)
                                )
                                connection.commit()
                                total_fixed += 1
                            except Exception as row_error:
                                logging.error(f"Error updating ResultID {result_id}: {row_error}")
                                error_count += 1
            
            logging.info(f"JSON fix operation completed. Fixed: {total_fixed}, Errors: {error_count}")
            
            cursor.close()
            connection.close()
            
            return {
                "status": "completed",
                "records_processed": len(rows),
                "records_fixed": total_fixed,
                "errors": error_count
            }
        except Exception as e:
            logging.error(f"Error in background JSON fix: {e}")
            logging.error(traceback.format_exc())
            return {
                "status": "error",
                "error_message": str(e)
            }
    
    background_tasks.add_task(background_fix_json)
    return {
        "message": f"JSON fix operation initiated in background" + (f" for FileID: {file_id}" if file_id else " across all files"),
        "status": "processing",
        "limit": limit
    }

def clean_json(value):
    if not value or not isinstance(value, str) or value.strip() in ["None", "null", "NaN", "undefined"]:
        return json.dumps({
            "description": "",
            "user_provided": {"brand": "", "category": "", "color": ""},
            "extracted_features": {"brand": "", "category": "", "color": ""},
            "match_score": None,
            "reasoning_match": "",
            "linesheet_score": None,
            "reasoning_linesheet": ""
        })
    
    value = value.strip()
    if value and value[0] not in ['{', '[', '"']:
        logging.warning(f"Invalid JSON starting with '{value[0:10]}...' - replacing with default")
        return json.dumps({
            "description": "",
            "user_provided": {"brand": "", "category": "", "color": ""},
            "extracted_features": {"brand": "", "category": "", "color": ""},
            "match_score": None,
            "reasoning_match": "",
            "linesheet_score": None,
            "reasoning_linesheet": ""
        })
    
    try:
        parsed = json.loads(value)
        if not isinstance(parsed, dict):
            logging.warning("JSON is not a dictionary - replacing with default structure")
            return json.dumps({
                "description": "",
                "user_provided": {"brand": "", "category": "", "color": ""},
                "extracted_features": {"brand": "", "category": "", "color": ""},
                "match_score": None,
                "reasoning_match": "",
                "linesheet_score": None,
                "reasoning_linesheet": ""
            })
        
        if "linesheet_score" in parsed:
            if isinstance(parsed["linesheet_score"], float) and math.isnan(parsed["linesheet_score"]):
                parsed["linesheet_score"] = None
            elif parsed["linesheet_score"] in ["NaN", "null", "undefined", ""]:
                parsed["linesheet_score"] = None
                
        if "match_score" in parsed:
            if isinstance(parsed["match_score"], float) and math.isnan(parsed["match_score"]):
                parsed["match_score"] = None
            elif parsed["match_score"] in ["NaN", "null", "undefined", ""]:
                parsed["match_score"] = None
        
        if "description" not in parsed:
            parsed["description"] = ""
        if "user_provided" not in parsed:
            parsed["user_provided"] = {"brand": "", "category": "", "color": ""}
        if "extracted_features" not in parsed:
            parsed["extracted_features"] = {"brand": "", "category": "", "color": ""}
        if "match_score" not in parsed:
            parsed["match_score"] = None
        if "reasoning_match" not in parsed:
            parsed["reasoning_match"] = ""
        if "linesheet_score" not in parsed:
            parsed["linesheet_score"] = None
        if "reasoning_linesheet" not in parsed:
            parsed["reasoning_linesheet"] = ""
            
        for field_dict in ["user_provided", "extracted_features"]:
            if field_dict in parsed and isinstance(parsed[field_dict], dict):
                for field in ["brand", "category", "color"]:
                    if field not in parsed[field_dict]:
                        parsed[field_dict][field] = ""

        return json.dumps(parsed)

    except json.JSONDecodeError as e:
        logging.warning(f"JSON decoding error: {e} for value: {value[:50]}...")
        return json.dumps({
            "description": "",
            "user_provided": {"brand": "", "category": "", "color": ""},
            "extracted_features": {"brand": "", "category": "", "color": ""},
            "match_score": None,
            "reasoning_match": "",
            "linesheet_score": None,
            "reasoning_linesheet": ""
        })
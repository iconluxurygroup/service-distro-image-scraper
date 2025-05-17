import logging
import pandas as pd
import asyncio
import json
import datetime
import os
import time
import hashlib
import psutil
import httpx
from fastapi import BackgroundTasks, HTTPException, APIRouter, Query
from pydantic import BaseModel, Field
from sqlalchemy.sql import text
from sqlalchemy.exc import SQLAlchemyError
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from database_config import conn_str, async_engine
from s3_utils import upload_file_to_space
from db_utils import (
    get_send_to_email,
    get_images_excel_db,
    update_file_location_complete,
    update_file_generate_complete,
    fetch_last_valid_entry,
    update_log_url_in_db,
)
from search_utils import update_search_sort_order, insert_search_results
from common import fetch_brand_rules, clean_string, generate_aliases, generate_brand_aliases
from utils import create_temp_dirs, cleanup_temp_dirs, process_and_tag_results
from endpoint_utils import sync_get_endpoint
from logging_config import setup_job_logger
from email_utils import send_message_email
from search_variations import generate_search_variations    
import aiofiles
import aiohttp
from typing import Optional, List, Dict, Tuple, Any, Callable
from urllib.parse import urlparse
from url_extract import extract_thumbnail_url
import re
from operator import itemgetter
from PIL import Image as IMG2
from io import BytesIO
import numpy as np
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

default_logger = logging.getLogger(__name__)
if not default_logger.handlers:
    default_logger.setLevel(logging.INFO)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

BRAND_RULES_URL = os.getenv("BRAND_RULES_URL", "https://raw.githubusercontent.com/iconluxurygroup/legacy-icon-product-api/refs/heads/main/task_settings/brand_settings.json")

JOB_STATUS = {}
LAST_UPLOAD = {}

class SearchClient:
    def __init__(self, endpoint: str, logger: logging.Logger, max_concurrency: int = 2):
        self.endpoint = endpoint
        self.logger = logger
        self.semaphore = asyncio.Semaphore(max_concurrency)
        self.client = httpx.AsyncClient(timeout=10.0)

    async def close(self):
        await self.client.aclose()

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        retry=retry_if_exception_type((httpx.HTTPStatusError, httpx.RequestError))
    )
    async def search(self, term: str, brand: str) -> List[Dict]:
        async with self.semaphore:
            try:
                response = await self.client.get(
                    self.endpoint,
                    params={"q": term, "brand": brand}
                )
                response.raise_for_status()
                return response.json().get("results", [])
            except httpx.HTTPStatusError as e:
                self.logger.error(f"HTTP error for term '{term}': {e}")
                raise
            except httpx.RequestError as e:
                self.logger.error(f"Request error for term '{term}': {e}")
                raise

async def process_results(
    raw_results: List[Dict],
    entry_id: int,
    brand: str,
    search_term: str,
    logger: logging.Logger
) -> List[Dict]:
    results = []
    required_columns = ["EntryID", "ImageUrl", "ImageDesc", "ImageSource", "ImageUrlThumbnail"]

    tagged_results = await process_and_tag_results(raw_results, brand=brand, search_term=search_term, logger=logger)
    for result in tagged_results:
        image_url = result.get("image_url")
        if not image_url:
            continue

        thumbnail_url = await extract_thumbnail_url(image_url, logger=logger) or image_url
        parsed_url = urlparse(image_url)
        image_source = parsed_url.netloc or "unknown"

        formatted_result = {
            "EntryID": entry_id,
            "ImageUrl": image_url,
            "ImageDesc": result.get("description", ""),
            "ImageSource": image_source,
            "ImageUrlThumbnail": thumbnail_url
        }

        if all(col in formatted_result for col in required_columns):
            results.append(formatted_result)
        else:
            logger.warning(f"Result missing required columns for EntryID {entry_id}: {formatted_result}")

    return results

async def async_process_entry_search(
    search_string: str,
    brand: str,
    endpoint: str,
    entry_id: int,
    use_all_variations: bool,
    file_id_db: int,
    logger: logging.Logger
) -> List[Dict]:
    logger.debug(f"Processing search for EntryID {entry_id}, FileID {file_id_db}, Use all variations: {use_all_variations}")
    search_terms_dict = generate_search_variations(search_string, brand, logger=logger)
    
    # Collect all search terms based on use_all_variations
    search_terms = []
    variation_types = [
        "default", "delimiter_variations", "color_variations", "brand_alias",
        "no_color", "model_alias", "category_specific"
    ]
    
    for variation_type in variation_types:
        if variation_type in search_terms_dict:
            search_terms.extend(search_terms_dict[variation_type])
            if not use_all_variations:
                break  # Stop after the first variation type unless all variations are requested
    
    search_terms = list(set(search_terms))  # Deduplicate terms
    logger.info(f"Generated {len(search_terms)} search terms for EntryID {entry_id}")

    if not search_terms:
        logger.warning(f"No search terms for EntryID {entry_id}")
        return []

    client = SearchClient(endpoint, logger)
    try:
        tasks = [client.search(term, brand) for term in search_terms]
        raw_results = await asyncio.gather(*tasks, return_exceptions=True)

        all_results = []
        for term, term_results in zip(search_terms, raw_results):
            if isinstance(term_results, Exception):
                logger.error(f"Error for term '{term}' in EntryID {entry_id}: {term_results}")
                continue
            if not term_results:
                continue
            results = await process_results(term_results, entry_id, brand, term, logger)
            all_results.extend(results)

        logger.info(f"Processed {len(all_results)} results for EntryID {entry_id}")
        return all_results
    finally:
        await client.close()

async def generate_download_file(file_id: int, background_tasks: BackgroundTasks, logger: Optional[logging.Logger] = None) -> Dict[str, str]:
    """
    Generate an Excel file containing image scraping results for the given file_id and upload it to R2.
    """
    log_filename = f"job_logs/job_{file_id}.log"
    try:
        if logger is None:
            logger, log_filename = setup_job_logger(job_id=str(file_id), log_dir="job_logs", console_output=True)
        logger.setLevel(logging.DEBUG)
        process = psutil.Process()
        logger.debug(f"Logger initialized for generate_download_file")

        def log_memory_usage():
            mem_info = process.memory_info()
            logger.info(f"Memory: RSS={mem_info.rss / 1024**2:.2f} MB")
            if mem_info.rss / 1024**2 > 1000:
                logger.warning(f"High memory usage")

        logger.info(f"Generating download file for FileID: {file_id}")
        log_memory_usage()

        # Fetch image scraping results
        results_df = await get_images_excel_db(str(file_id), logger)
        if results_df.empty:
            logger.error(f"No data found for FileID {file_id}")
            background_tasks.add_task(monitor_and_resubmit_failed_jobs, str(file_id), logger)
            return {"error": f"No data found for FileID {file_id}", "log_filename": log_filename}

        # Create temporary directory for Excel file
        temp_dir = f"temp_excel_{file_id}"
        os.makedirs(temp_dir, exist_ok=True)
        excel_filename = os.path.join(temp_dir, f"image_results_{file_id}.xlsx")

        # Generate Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Results"

        # Define headers
        headers = [
            "EntryID", "ProductBrand", "ProductModel", "ProductColor", "ProductCategory",
            "ImageUrl", "ImageDesc", "ImageSource", "ImageUrlThumbnail", "SortOrder"
        ]
        for col, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col)}1"] = header
            ws[f"{get_column_letter(col)}1"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write data
        for row_idx, row in results_df.iterrows():
            for col_idx, header in enumerate(headers):
                ws[f"{get_column_letter(col_idx + 1)}{row_idx + 2}"] = row.get(header, "")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_filename)
        logger.info(f"Excel file generated: {excel_filename}")

        # Upload to R2
        public_url = await upload_file_to_space(
            file_src=excel_filename,
            save_as=f"excel_results/image_results_{file_id}.xlsx",
            is_public=True,
            logger=logger,
            file_id=str(file_id)
        )
        if not public_url:
            logger.error(f"Failed to upload Excel file for FileID {file_id}")
            return {"error": "Failed to upload Excel file", "log_filename": log_filename}

        # Update database
        await update_file_location_complete(str(file_id), public_url, logger)
        await update_file_generate_complete(str(file_id), logger)

        # Clean up temporary directory
        try:
            os.remove(excel_filename)
            os.rmdir(temp_dir)
            logger.debug(f"Cleaned up temporary directory: {temp_dir}")
        except Exception as e:
            logger.warning(f"Failed to clean up temporary directory {temp_dir}: {e}")

        logger.info(f"Download file generated and uploaded for FileID: {file_id}: {public_url}")
        return {
            "message": "Download file generated successfully",
            "file_id": str(file_id),
            "public_url": public_url,
            "log_filename": log_filename
        }
    except Exception as e:
        logger.error(f"Error generating download file for FileID {file_id}: {e}", exc_info=True)
        background_tasks.add_task(monitor_and_resubmit_failed_jobs, str(file_id), logger)
        return {"error": str(e), "log_filename": log_filename}
    finally:
        log_memory_usage()

async def run_generate_download_file(file_id: str, logger: logging.Logger, log_filename: str, background_tasks: BackgroundTasks):
    try:
        JOB_STATUS[file_id] = {
            "status": "running",
            "message": "Job is running",
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        result = await generate_download_file(int(file_id), background_tasks, logger=logger)
        
        if "error" in result:
            JOB_STATUS[file_id] = {
                "status": "failed",
                "message": f"Error: {result['error']}",
                "log_url": result.get("log_filename") if os.path.exists(result.get("log_filename", "")) else None,
                "timestamp": datetime.datetime.now().isoformat()
            }
            logger.error(f"Job failed for FileID {file_id}: {result['error']}")
        else:
            JOB_STATUS[file_id] = {
                "status": "completed",
                "message": "Job completed successfully",
                "public_url": result.get("public_url"),
                "log_url": result.get("log_filename") if os.path.exists(result.get("log_filename", "")) else None,
                "timestamp": datetime.datetime.now().isoformat()
            }
            logger.info(f"Job completed for FileID {file_id}")
    except Exception as e:
        logger.error(f"Unexpected error in job for FileID {file_id}: {e}", exc_info=True)
        log_public_url = await upload_log_file(file_id, log_filename, logger)
        JOB_STATUS[file_id] = {
            "status": "failed",
            "message": f"Unexpected error: {str(e)}",
            "log_url": log_public_url or None,
            "timestamp": datetime.datetime.now().isoformat()
        }
        background_tasks.add_task(monitor_and_resubmit_failed_jobs, file_id, logger)

async def monitor_and_resubmit_failed_jobs(file_id: str, logger: logging.Logger):
    log_file = f"job_logs/job_{file_id}.log"
    max_attempts = 3
    attempt = 1

    while attempt <= max_attempts:
        if os.path.exists(log_file):
            with open(log_file, 'r') as f:
                log_content = f.read()
                if any(err in log_content for err in ["WORKER TIMEOUT", "SIGKILL", "placeholder://error", "No data found"]):
                    logger.warning(f"Detected failure in job for FileID: {file_id}, attempt {attempt}/{max_attempts}")
                    last_entry_id = await fetch_last_valid_entry(file_id, logger)
                    logger.info(f"Resubmitting job for FileID: {file_id} starting from EntryID: {last_entry_id or 'beginning'} with all variations")
                    
                    result = await process_restart_batch(
                        file_id_db=int(file_id),
                        logger=logger,
                        entry_id=last_entry_id,
                        use_all_variations=True
                    )
                    
                    if "error" not in result:
                        logger.info(f"Resubmission successful for FileID: {file_id}")
                        # Retry generating download file
                        await run_generate_download_file(file_id, logger, log_file, BackgroundTasks())
                        await send_message_email(
                            to_emails=["nik@luxurymarket.com"],
                            subject=f"Success: Batch Resubmission for FileID {file_id}",
                            message=f"Resubmission succeeded for FileID {file_id} starting from EntryID {last_entry_id or 'beginning'} with all variations.\nLog: {log_file}",
                            logger=logger
                        )
                        return
                    else:
                        logger.error(f"Resubmission failed for FileID: {file_id}: {result['error']}")
                    attempt += 1
                    await asyncio.sleep(2 ** attempt)  # Exponential backoff
                else:
                    logger.info(f"No failure detected in logs for FileID: {file_id}")
                    return
        else:
            logger.warning(f"Log file {log_file} does not exist for FileID: {file_id}")
            return
        await asyncio.sleep(60)

router = APIRouter()

class JobStatusResponse(BaseModel):
    status: str = Field(..., description="Job status (e.g., queued, running, completed, failed)")
    message: str = Field(..., description="Descriptive message about the job status")
    public_url: Optional[str] = Field(None, description="R2 URL of the generated Excel file, if available")
    log_url: Optional[str] = Field(None, description="R2 URL of the job log file, if available")
    timestamp: str = Field(..., description="ISO timestamp of the response")

@router.post("/generate-download-file/{file_id}", tags=["Export"], response_model=JobStatusResponse)
async def api_generate_download_file(file_id: str, background_tasks: BackgroundTasks):
    logger, log_filename = setup_job_logger(job_id=file_id, console_output=True)
    logger.info(f"Received request to generate download file for FileID: {file_id}")
    
    try:
        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("SELECT FileName FROM utb_ImageScraperFiles WHERE ID = :file_id"),
                {"file_id": int(file_id)}
            )
            if not result.fetchone():
                logger.error(f"Invalid FileID: {file_id}")
                raise HTTPException(status_code=404, detail=f"FileID {file_id} not found")
        
        JOB_STATUS[file_id] = {
            "status": "queued",
            "message": "Job queued for processing",
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        background_tasks.add_task(run_generate_download_file, file_id, logger, log_filename, background_tasks)
        
        send_to_email = await get_send_to_email(int(file_id), logger=logger)
        if send_to_email:
            await send_message_email(
                to_emails=send_to_email,
                subject=f"Job Queued for FileID: {file_id}",
                message=f"Excel file generation for FileID {file_id} has been queued.",
                logger=logger
            )
        
        return JobStatusResponse(
            status="queued",
            message=f"Download file generation queued for FileID: {file_id}",
            timestamp=datetime.datetime.now().isoformat()
        )
    except SQLAlchemyError as e:
        logger.error(f"Database error for FileID {file_id}: {e}", exc_info=True)
        log_public_url = await upload_log_file(file_id, log_filename, logger)
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        logger.error(f"Error queuing download file for FileID {file_id}: {e}", exc_info=True)
        log_public_url = await upload_log_file(file_id, log_filename, logger)
        raise HTTPException(status_code=500, detail=f"Error queuing job: {str(e)}")

async def process_restart_batch(
    file_id_db: int,
    entry_id: Optional[int] = None,
    use_all_variations: bool = False,
    logger: Optional[logging.Logger] = None
) -> Dict[str, str]:
    log_filename = f"job_logs/job_{file_id_db}.log"
    try:
        if logger is None:
            logger, log_filename = setup_job_logger(job_id=str(file_id_db), log_dir="job_logs", console_output=True)
        logger.setLevel(logging.DEBUG)
        process = psutil.Process()
        logger.debug(f"Logger initialized")

        def log_memory_usage():
            mem_info = process.memory_info()
            logger.info(f"Memory: RSS={mem_info.rss / 1024**2:.2f} MB")
            if mem_info.rss / 1024**2 > 1000:
                logger.warning(f"High memory usage")

        logger.info(f"Starting processing for FileID: {file_id_db}, Use all variations: {use_all_variations}")
        log_memory_usage()

        file_id_db_int = file_id_db
        BATCH_SIZE = 1
        MAX_CONCURRENCY = 4
        MAX_ENTRY_RETRIES = 3

        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = :file_id"),
                {"file_id": file_id_db_int}
            )
            if result.fetchone()[0] == 0:
                logger.error(f"FileID {file_id_db} does not exist")
                return {"error": f"FileID {file_id_db} does not exist", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}
            result.close()

        if entry_id is None:
            entry_id = await fetch_last_valid_entry(str(file_id_db_int), logger)
            if entry_id is not None:
                async with async_engine.connect() as conn:
                    result = await conn.execute(
                        text("SELECT MIN(EntryID) FROM utb_ImageScraperRecords WHERE FileID = :file_id AND EntryID > :entry_id"),
                        {"file_id": file_id_db_int, "entry_id": entry_id}
                    )
                    next_entry = result.fetchone()
                    entry_id = next_entry[0] if next_entry and next_entry[0] else None
                    logger.info(f"Resuming from EntryID: {entry_id}")
                    result.close()

        brand_rules = await fetch_brand_rules(BRAND_RULES_URL, max_attempts=3, timeout=10, logger=logger)
        if not brand_rules:
            logger.warning(f"No brand rules fetched")
            return {"message": "Failed to fetch brand rules", "file_id": str(file_id_db), "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        endpoint = None
        for attempt in range(5):
            try:
                endpoint = sync_get_endpoint(logger=logger)
                if endpoint:
                    logger.info(f"Selected endpoint: {endpoint}")
                    break
                logger.warning(f"Attempt {attempt + 1} failed")
                await asyncio.sleep(2)
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed: {e}")
                await asyncio.sleep(2)
        if not endpoint:
            logger.error(f"No healthy endpoint")
            return {"error": "No healthy endpoint", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        async with async_engine.connect() as conn:
            query = text("""
                SELECT EntryID, ProductModel, ProductBrand, ProductColor, ProductCategory 
                FROM utb_ImageScraperRecords 
                WHERE FileID = :file_id AND (:entry_id IS NULL OR EntryID >= :entry_id) 
                ORDER BY EntryID
            """)
            result = await conn.execute(query, {"file_id": file_id_db_int, "entry_id": entry_id})
            entries = [(row[0], row[1], row[2], row[3], row[4]) for row in result.fetchall() if row[1] is not None]
            logger.info(f"Found {len(entries)} entries")
            result.close()

        if not entries:
            logger.warning(f"No valid EntryIDs found for FileID {file_id_db}")
            return {"error": "No entries found", "log_filename": log_filename, "log_public_url": "", "last_entry_id": str(entry_id or "")}

        entry_batches = [entries[i:i + BATCH_SIZE] for i in range(0, len(entries), BATCH_SIZE)]
        logger.info(f"Created {len(entry_batches)} batches")

        successful_entries = 0
        failed_entries = 0
        last_entry_id_processed = entry_id or 0
        required_columns = ["EntryID", "ImageUrl", "ImageDesc", "ImageSource", "ImageUrlThumbnail"]

        semaphore = asyncio.Semaphore(MAX_CONCURRENCY)
        async def process_entry(entry):
            entry_id, search_string, brand, color, category = entry
            async with semaphore:
                for attempt in range(1, MAX_ENTRY_RETRIES + 1):
                    try:
                        logger.info(f"Processing EntryID {entry_id}, Attempt {attempt}/{MAX_ENTRY_RETRIES}, Use all variations: {use_all_variations}")
                        results = await async_process_entry_search(
                            search_string=search_string,
                            brand=brand,
                            endpoint=endpoint,
                            entry_id=entry_id,
                            use_all_variations=use_all_variations,
                            file_id_db=file_id_db,
                            logger=logger
                        )
                        if not results:
                            logger.warning(f"No results for EntryID {entry_id} on attempt {attempt}")
                            continue

                        if not all(all(col in res for col in required_columns) for res in results):
                            logger.error(f"Missing columns for EntryID {entry_id} on attempt {attempt}")
                            continue

                        deduplicated_results = []
                        seen = set()
                        for res in results:
                            key = (res['EntryID'], res['ImageUrl'])
                            if key not in seen:
                                seen.add(key)
                                deduplicated_results.append(res)
                        logger.info(f"Deduplicated to {len(deduplicated_results)} rows for EntryID {entry_id}")

                        insert_success = await insert_search_results(deduplicated_results, logger=logger, file_id=str(file_id_db))
                        if not insert_success:
                            logger.error(f"Failed to insert results for EntryID {entry_id} on attempt {attempt}")
                            continue

                        update_result = await update_search_sort_order(
                            str(file_id_db), str(entry_id), brand, search_string, color, category, logger, brand_rules=brand_rules
                        )
                        if update_result is None or not update_result:
                            logger.error(f"SortOrder update failed for EntryID {entry_id} on attempt {attempt}")
                            continue

                        return entry_id, True
                    except Exception as e:
                        logger.error(f"Error processing EntryID {entry_id} on attempt {attempt}: {e}", exc_info=True)
                        if attempt < MAX_ENTRY_RETRIES:
                            await asyncio.sleep(2 ** attempt)  # Exponential backoff
                        continue
                logger.error(f"Failed to process EntryID {entry_id} after {MAX_ENTRY_RETRIES} attempts")
                return entry_id, False

        for batch_idx, batch_entries in enumerate(entry_batches, 1):
            logger.info(f"Processing batch {batch_idx}/{len(entry_batches)}")
            start_time = datetime.datetime.now()

            results = await asyncio.gather(
                *(process_entry(entry) for entry in batch_entries),
                return_exceptions=True
            )

            for entry, result in zip(batch_entries, results):
                entry_id = entry[0]
                if isinstance(result, Exception):
                    logger.error(f"Error processing EntryID {entry_id}: {result}", exc_info=True)
                    failed_entries += 1
                    continue
                entry_id_result, success = result
                if success:
                    successful_entries += 1
                    last_entry_id_processed = entry_id
                else:
                    failed_entries += 1

            elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
            logger.info(f"Completed batch {batch_idx} in {elapsed_time:.2f}s")
            log_memory_usage()

        async with async_engine.connect() as conn:
            result = await conn.execute(
                text("""
                    SELECT COUNT(DISTINCT t.EntryID), 
                           SUM(CASE WHEN t.SortOrder > 0 THEN 1 ELSE 0 END) AS positive_count,
                           SUM(CASE WHEN t.SortOrder IS NULL THEN 1 ELSE 0 END) AS null_count
                    FROM utb_ImageScraperResult t
                    INNER JOIN utb_ImageScraperRecords r ON t.EntryID = r.EntryID
                    WHERE r.FileID = :file_id
                """),
                {"file_id": file_id_db_int}
            )
            row = result.fetchone()
            total_entries = row[0] if row else 0
            positive_entries = row[1] if row and row[1] is not None else 0
            null_entries = row[2] if row and row[2] is not None else 0
            logger.info(
                f"Verification: {total_entries} total entries, "
                f"{positive_entries} with positive SortOrder, {null_entries} with NULL SortOrder"
            )
            if null_entries > 0:
                logger.warning(f"Found {null_entries} entries with NULL SortOrder")
            result.close()

        to_emails = await get_send_to_email(file_id_db, logger=logger)
        if to_emails:
            subject = f"Processing Completed for FileID: {file_id_db}"
            message = (
                f"Processing for FileID {file_id_db} completed.\n"
                f"Successful entries: {successful_entries}/{len(entries)}\n"
                f"Failed entries: {failed_entries}\n"
                f"Last EntryID: {last_entry_id_processed}\n"
                f"Log file: {log_filename}\n"
                f"Used all variations: {use_all_variations}"
            )
            await send_message_email(to_emails, subject=subject, message=message, logger=logger)

        log_public_url = await upload_log_file(str(file_id_db), log_filename, logger)
        return {
            "message": "Search processing completed",
            "file_id": str(file_id_db),
            "successful_entries": str(successful_entries),
            "total_entries": str(len(entries)),
            "failed_entries": str(failed_entries),
            "log_filename": log_filename,
            "log_public_url": log_public_url or "",
            "last_entry_id": str(last_entry_id_processed),
            "use_all_variations": str(use_all_variations)
        }
    except Exception as e:
        logger.error(f"Error processing FileID {file_id_db}: {e}", exc_info=True)
        log_public_url = await upload_log_file(str(file_id_db), log_filename, logger)
        return {"error": str(e), "log_filename": log_filename, "log_public_url": log_public_url or "", "last_entry_id": str(entry_id or "")}
    finally:
        await async_engine.dispose()
        logger.info(f"Disposed database engines")

async def upload_log_file(file_id: str, log_filename: str, logger: logging.Logger) -> Optional[str]:
    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        retry=retry_if_exception_type(Exception),
        before_sleep=lambda retry_state: logger.info(
            f"Retrying log upload for FileID {file_id} (attempt {retry_state.attempt_number}/3)"
        )
    )
    async def try_upload():
        if not os.path.exists(log_filename):
            logger.warning(f"Log file {log_filename} does not exist, skipping upload")
            return None

        file_hash = await asyncio.to_thread(hashlib.md5, open(log_filename, "rb").read())
        file_hash = file_hash.hexdigest()
        current_time = time.time()
        key = (log_filename, file_id)

        if key in LAST_UPLOAD and LAST_UPLOAD[key]["hash"] == file_hash and current_time - LAST_UPLOAD[key]["time"] < 60:
            logger.info(f"Skipping redundant upload for {log_filename}")
            return LAST_UPLOAD[key]["url"]

        try:
            upload_url = await upload_file_to_space(
                file_src=log_filename,
                save_as=f"job_logs/job_{file_id}.log",
                is_public=True,
                logger=logger,
                file_id=file_id
            )
            if not upload_url:
                logger.error(f"S3 upload returned empty URL for {log_filename}")
                raise ValueError("Empty upload URL")
            await update_log_url_in_db(file_id, upload_url, logger)
            LAST_UPLOAD[key] = {"hash": file_hash, "time": current_time, "url": upload_url}
            logger.info(f"Log uploaded to R2: {upload_url}")
            return upload_url
        except Exception as e:
            logger.error(f"Failed to upload log for FileID {file_id}: {e}", exc_info=True)
            raise

    try:
        return await try_upload()
    except Exception as e:
        logger.error(f"Failed to upload log for FileID {file_id} after retries: {e}", exc_info=True)
        return None
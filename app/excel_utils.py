import os
import logging
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict
from PIL import Image as IMG2
import numpy as np
from collections import Counter
import urllib.parse
import re

# Module-level logger
default_logger = logging.getLogger(__name__)
if not default_logger.handlers:
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

def clean_string(value: str, is_url: bool = False) -> str:
    """Clean a string by removing backslashes and decoding URL-encoded characters."""
    if not value:
        return ""
    cleaned = str(value).replace('\\', '').replace('%5C', '').replace('%5c', '')
    cleaned = re.sub(r'[\x00-\x1F\x7F]+', '', cleaned).strip()
    if is_url:
        cleaned = urllib.parse.unquote(cleaned)
        try:
            parsed = urllib.parse.urlparse(cleaned)
            if parsed.scheme and parsed.netloc:
                path = re.sub(r'/+', '/', parsed.path)
                cleaned = f"{parsed.scheme}://{parsed.netloc}{path}"
                if parsed.query:
                    cleaned += f"?{parsed.query}"
                if parsed.fragment:
                    cleaned += f"#{parsed.fragment}"
        except ValueError:
            cleaned = ""
    return cleaned

def verify_png_image_single(image_path, logger=None):
    logger = logger or default_logger
    try:
        logger.debug(f"🔎 Verifying image: {image_path}")
        img = IMG2.open(image_path)
        img.verify()
        logger.info(f"✅ Image verified successfully: {image_path}")
    except Exception as e:
        logger.error(f"❌ Image verify failed: {e}, for image: {image_path}", exc_info=True)
        return False

    imageSize = os.path.getsize(image_path)
    logger.debug(f"📏 Image size: {imageSize} bytes")

    if imageSize < 3000:
        logger.warning(f"⚠️ File may be corrupted or too small: {image_path}")
        return False

    try:
        if not resize_image(image_path, logger=logger):
            logger.warning(f"⚠️ Resize failed for: {image_path}")
            return False
        img = IMG2.open(image_path)
        img.verify()
        logger.info(f"✅ Post-resize verification successful: {image_path}")
        return True
    except Exception as e:
        logger.error(f"❌ Error during verification or resizing: {e}, for image: {image_path}", exc_info=True)
        return False

def resize_image(image_path, logger=None):
    logger = logger or default_logger
    try:
        logger.debug(f"📂 Attempting to open image: {image_path}")
        img = IMG2.open(image_path)
        MAXSIZE = 130

        if img.mode == 'RGBA':
            logger.info(f"🌈 Converting RGBA image to RGB with white background: {image_path}")
            background = IMG2.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])
            img = background
        elif img.mode != 'RGB':
            logger.info(f"🌈 Converting {img.mode} image to RGB: {image_path}")
            img = img.convert('RGB')

        def get_background_color(img):
            width, height = img.size
            pixels = np.array(img)
            top = pixels[0, :]
            bottom = pixels[height-1, :]
            left = pixels[1:height-1, 0]
            right = pixels[1:height-1, width-1]
            border_pixels = np.concatenate((top, bottom, left, right))
            color_counts = Counter(map(tuple, border_pixels))
            most_common_color, count = color_counts.most_common(1)[0]
            total_border_pixels = border_pixels.shape[0]
            if count / total_border_pixels >= 0.9:
                return most_common_color
            return None

        def is_white(color, threshold=240):
            r, g, b = color
            return r >= threshold and g >= threshold and b >= threshold

        def replace_background(img, bg_color):
            pixels = np.array(img)
            bg_color = np.array(bg_color)
            white = np.array([255, 255, 255])
            diff = np.abs(pixels - bg_color)
            mask = np.all(diff <= 5, axis=2)
            pixels[mask] = white
            return IMG2.fromarray(pixels)

        background_color = get_background_color(img)
        if background_color and not is_white(background_color):
            logger.info(f"🖌️ Replacing background color {background_color} with white for {image_path}")
            img = replace_background(img, background_color)

        h, w = img.height, img.width
        logger.debug(f"📐 Original size: height={h}, width={w}")
        if h > MAXSIZE or w > MAXSIZE:
            if h > w:
                w = int(w * MAXSIZE / h)
                h = MAXSIZE
            else:
                h = int(h * MAXSIZE / w)
                w = MAXSIZE
            logger.debug(f"🔍 Resizing to: height={h}, width={w}")
            newImg = img.resize((w, h))
        else:
            newImg = img

        newImg.save(image_path, 'PNG')
        logger.info(f"✅ Image processed and saved: {image_path}")
        if os.path.exists(image_path):
            logger.debug(f"📏 File size after save: {os.path.getsize(image_path)} bytes")
        else:
            logger.error(f"❌ File not found after save: {image_path}")
            return False
        return True

    except Exception as e:
        logger.error(f"❌ Error resizing image: {e}, for image: {image_path}", exc_info=True)
        return False

def write_excel_image(local_filename, temp_dir, image_data: List[Dict], column="A", row_offset=5, logger=None):
    """Write one image per entry to an Excel file starting at row 6, removing unneeded rows, trying multiple sort orders."""
    logger = logger or default_logger
    failed_rows = []

    try:
        logger.debug(f"📂 Loading workbook from {local_filename}")
        wb = load_workbook(local_filename)
        ws = wb.active

        if ws.max_row < 5:
            logger.error(f"❌ Excel file must have at least 5 rows for header")
            return failed_rows

        logger.info(f"🖼️ Processing images for {local_filename}")
        if not os.path.exists(temp_dir):
            logger.error(f"❌ Temp directory does not exist: {temp_dir}")
            return failed_rows

        # Clean and list image files
        image_files = [clean_string(f, is_url=False) for f in os.listdir(temp_dir)]
        if not image_files:
            logger.warning(f"⚠️ No images found in {temp_dir}")
            return failed_rows

        # Clean and validate image_data
        cleaned_data = []
        for item in image_data:
            cleaned_item = {
                'ExcelRowID': item.get('ExcelRowID', 0),
                'Brand': clean_string(item.get('Brand', '')),
                'Style': clean_string(item.get('Style', '')),
                'Color': clean_string(item.get('Color', '')),
                'Category': clean_string(item.get('Category', '')),
                'SortOrder': item.get('SortOrder', 9999)
            }
            if not isinstance(cleaned_item['ExcelRowID'], int) or cleaned_item['ExcelRowID'] < 1:
                logger.warning(f"⚠️ Invalid ExcelRowID: {cleaned_item['ExcelRowID']}")
                continue
            cleaned_data.append(cleaned_item)
            logger.debug(f"🧹 Cleaned item: {cleaned_item}")
        logger.info(f"🧹 Cleaned {len(cleaned_data)} of {len(image_data)} entries")

        # Create image map
        image_map = {}
        for f in image_files:
            if '_' in f and f.split('_')[0].isdigit():
                row_id = int(f.split('_')[0])
                image_map[row_id] = f

        # Define sorting strategies
        sort_strategies = [
            ('SortOrder', lambda x: (x['SortOrder'], -1 if x['ExcelRowID'] >= 2 else 0, x['ExcelRowID'])),
            ('ImageAvailability', lambda x: (0 if x['ExcelRowID'] in image_map else 1, -1 if x['ExcelRowID'] >= 2 else 0, x['ExcelRowID'])),
            ('Brand', lambda x: (x['Brand'].lower(), -1 if x['ExcelRowID'] >= 2 else 0, x['ExcelRowID'])),
            ('ExcelRowID', lambda x: x['ExcelRowID'])
        ]

        best_failed_rows = []
        min_failures = float('inf')
        best_sorted_data = cleaned_data

        # Try each sort strategy
        for strategy_name, sort_key in sort_strategies:
            failed_rows.clear()
            logger.info(f"📊 Trying sort strategy: {strategy_name}")
            sorted_data = sorted(cleaned_data, key=sort_key)
            logger.debug(f"📊 Sorted order: {[item['ExcelRowID'] for item in sorted_data]}")

            # Process entries
            for item in sorted_data:
                row_id = item['ExcelRowID']
                row_number = row_id + row_offset

                success = False
                if row_id in image_map:
                    image_file = image_map[row_id]
                    image_path = os.path.join(temp_dir, image_file)
                    for attempt in range(2):  # Try twice
                        if verify_png_image_single(image_path, logger=logger):
                            img = Image(image_path)
                            anchor = f"{column}{row_number}"
                            img.anchor = anchor
                            ws.add_image(img)
                            logger.info(f"✅ Image added at {anchor} (attempt {attempt+1})")
                            success = True
                            break
                        logger.warning(f"⚠️ Image verification failed for {image_file} (attempt {attempt+1})")
                        # Retry with cleaned filename
                        image_path = os.path.join(temp_dir, clean_string(image_file, is_url=False))

                if not success:
                    logger.warning(f"⚠️ No image found or verification failed for row {row_id}")
                    failed_rows.append(row_id)

                # Populate cleaned data
                ws[f"B{row_number}"] = item['Brand']
                ws[f"D{row_number}"] = item['Style']
                ws[f"E{row_number}"] = item['Color']
                ws[f"H{row_number}"] = item['Category']
                logger.debug(f"✍️ Wrote data for row {row_number}: {item}")

                # Validate
                if not ws[f"B{row_number}"].value:
                    logger.warning(f"⚠️ Missing Brand in B{row_number}")
                if not ws[f"D{row_number}"].value:
                    logger.warning(f"⚠️ Missing Style in D{row_number}")

            logger.info(f"📊 Sort strategy {strategy_name} completed with {len(failed_rows)} failures")
            if len(failed_rows) < min_failures:
                min_failures = len(failed_rows)
                best_failed_rows = failed_rows.copy()
                best_sorted_data = sorted_data
                wb.save(local_filename + f".{strategy_name}.xlsx")  # Save for debugging
            if min_failures == 0:
                break

        # Use best sorted data for final output
        failed_rows = best_failed_rows
        for item in best_sorted_data:
            row_id = item['ExcelRowID']
            row_number = row_id + row_offset
            ws[f"B{row_number}"] = item['Brand']
            ws[f"D{row_number}"] = item['Style']
            ws[f"E{row_number}"] = item['Color']
            ws[f"H{row_number}"] = item['Category']

        # Remove unneeded rows
        max_data_row = max(item['ExcelRowID'] for item in best_sorted_data) + row_offset
        if ws.max_row > max_data_row:
            logger.info(f"🗑️ Removing rows {max_data_row + 1} to {ws.max_row}")
            ws.delete_rows(max_data_row + 1, ws.max_row - max_data_row)

        wb.save(local_filename)
        logger.info(f"🏁 Finished processing images with best strategy (failures: {len(failed_rows)})")
        return failed_rows

    except Exception as e:
        logger.error(f"❌ Error writing images: {e}", exc_info=True)
        return failed_rows

def highlight_cell(excel_file, cell_reference, logger=None):
    logger = logger or default_logger
    try:
        if not os.path.exists(excel_file):
            logger.error(f"❌ Excel file does not exist: {excel_file}")
            raise FileNotFoundError(f"Excel file not found: {excel_file}")
        
        logger.debug(f"📂 Loading workbook from {excel_file}")
        wb = load_workbook(excel_file)
        ws = wb.active
        
        if cell_reference not in ws:
            logger.error(f"❌ Invalid cell reference: {cell_reference} not in worksheet")
            raise ValueError(f"Invalid cell reference: {cell_reference}")
        
        logger.debug(f"🖌️ Applying yellow fill to cell {cell_reference}")
        ws[cell_reference].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        logger.debug(f"💾 Saving workbook to {excel_file}")
        wb.save(excel_file)
        
        wb_verify = load_workbook(excel_file)
        ws_verify = wb_verify.active
        fill_color = ws_verify[cell_reference].fill.start_color.index if ws_verify[cell_reference].fill else None
        if fill_color != "FFFF00":
            logger.warning(f"⚠️ Highlight not applied to {cell_reference} after save, found color: {fill_color}")
        else:
            logger.info(f"✅ Successfully highlighted cell {cell_reference} in {excel_file} with yellow fill")
        
    except FileNotFoundError as e:
        logger.error(f"❌ File error: {e}", exc_info=True)
        raise
    except ValueError as e:
        logger.error(f"❌ Value error: {e}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"❌ Error highlighting cell {cell_reference} in {excel_file}: {e}", exc_info=True)
        raise

def write_failed_downloads_to_excel(failed_downloads, excel_file, logger=None):
    logger = logger or default_logger
    if failed_downloads:
        try:
            logger.debug(f"📂 Loading workbook from {excel_file}")
            wb = load_workbook(excel_file)
            ws = wb.active
            
            max_row = ws.max_row
            max_failed_row = max(row_id for _, row_id in failed_downloads if row_id)
            if max_failed_row > max_row:
                logger.debug(f"Extending worksheet from {max_row} to {max_failed_row} rows")
                for i in range(max_row + 1, max_failed_row + 1):
                    ws[f"A{i}"] = ""
            
            for url, row_id in failed_downloads:
                if url and url != 'None found in this filter':
                    cleaned_url = clean_string(url, is_url=True)
                    cell_reference = f"{get_column_letter(1)}{row_id}"
                    logger.debug(f"✍️ Writing URL {cleaned_url} to cell {cell_reference}")
                    ws[cell_reference] = cleaned_url
                    for attempt in range(2):
                        try:
                            highlight_cell(excel_file, cell_reference, logger=logger)
                            break
                        except ValueError as e:
                            logger.warning(f"⚠️ Highlight attempt {attempt+1} failed for {cell_reference}: {e}")
                            continue
            
            logger.debug(f"💾 Saving workbook to {excel_file}")
            wb.save(excel_file)
            logger.info(f"✅ Failed downloads written to Excel file: {excel_file}")
            return True
        except Exception as e:
            logger.error(f"❌ Error writing failed downloads to Excel: {e}", exc_info=True)
            return False
    else:
        logger.info("ℹ️ No failed downloads to write to Excel.")
        return True
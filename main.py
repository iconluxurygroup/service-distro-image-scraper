import asyncio
import logging
import os
import shutil
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from typing import Optional, List, Tuple, Dict
from uuid import uuid4
import datetime
import aiohttp
import pandas as pd
import pyodbc
import requests
from aiohttp import ClientTimeout
from aiohttp_retry import RetryClient, ExponentialRetry
from fastapi import FastAPI, BackgroundTasks
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from PIL import UnidentifiedImageError
from tldextract import tldextract
import urllib.parse
from pathlib import Path
import numpy as np
from collections import Counter
import re
import random
from functools import partial

# --- Production Imports ---
# These are assumed to be in your project structure
from app_config import engine, conn_str
from email_utils import send_email, send_message_email
from s3_utils import upload_file_to_space

# --- Global Setup ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
app = FastAPI()

# --- Constants ---
MAX_THREADS = int(os.environ.get('MAX_THREADS', 10))
VALID_IMAGE_TYPES = ['image/jpeg', 'image/png', 'image/gif', 'image/bmp', 'image/webp', 'image/avif', 'image/tiff', 'image/x-icon']
MIN_IMAGE_SIZE = 1000  # Minimum content length in bytes
MAX_IMAGE_VERIFY_SIZE = 1000 # For verification before processing
MAX_IMAGE_DIMENSION = 130   # For resizing

def get_user_agents(logger_instance: logging.Logger) -> List[str]:
    """
    Fetches a list of user agents from the DataProxy API.
    Returns a default list if the API call fails.
    """
    default_ua = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"]
    try:
        api_url = "https://api.thedataproxy.com/v2/user-agents/?skip=0&limit=100"
        logger_instance.info(f"Fetching user agents from {api_url}")
        response = requests.get(api_url, timeout=15)
        response.raise_for_status()
        data = response.json()
        user_agents = [ua.get("user_agent") for ua in data.get("data", []) if ua.get("user_agent")]
        if not user_agents:
            logger_instance.warning("User-Agent API returned no agents. Using default.")
            return default_ua
        logger_instance.info(f"Successfully fetched {len(user_agents)} user agents.")
        return user_agents
    except requests.exceptions.RequestException as e:
        logger_instance.error(f"Failed to fetch user agents from API: {e}. Using default.")
        return default_ua

# --- General Utility Functions ---
def setup_logging(file_id: str, timestamp: str) -> logging.Logger:
    """Configure logging to save logs to a file for a specific job run."""
    log_dir = Path('jobs') / file_id / timestamp / 'logs'
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f'job_{file_id}_{timestamp}.log'
    
    logger_instance = logging.getLogger(f"job_{file_id}_{timestamp}")
    logger_instance.setLevel(logging.INFO)
    
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # Avoid adding handlers if they already exist
    if not logger_instance.handlers:
        file_handler = logging.FileHandler(log_file)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        
        logger_instance.addHandler(file_handler)
        logger_instance.addHandler(console_handler)
        logger_instance.propagate = False
    
    logger_instance.info(f"Logging initialized for FileID: {file_id}, Timestamp: {timestamp}")
    return logger_instance

async def create_temp_dirs(unique_id: str) -> Tuple[str, str]:
    """Create temporary directories for a job."""
    base_dir = Path.cwd() / 'temp_files'
    temp_images_dir = base_dir / 'images' / unique_id
    temp_excel_dir = base_dir / 'excel' / unique_id
    
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, partial(os.makedirs, temp_images_dir, exist_ok=True))
    await loop.run_in_executor(None, partial(os.makedirs, temp_excel_dir, exist_ok=True))
    
    return str(temp_images_dir), str(temp_excel_dir)

async def cleanup_temp_dirs(directories: List[str]):
    """Remove temporary directories."""
    loop = asyncio.get_running_loop()
    for dir_path in directories:
        await loop.run_in_executor(None, partial(shutil.rmtree, dir_path, ignore_errors=True))

# --- Database Functions ---
def get_file_type_id(file_id: int, logger_instance: logging.Logger) -> Optional[int]:
    """Retrieve the FileTypeID from the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "SELECT FileTypeID FROM utb_ImageScraperFiles WHERE ID = ?"
        cursor.execute(query, (file_id,))
        result = cursor.fetchone()
        logger_instance.info(f"FileTypeID for {file_id} is {result[0] if result else 'Not Found'}")
        return result[0] if result else None

def get_file_location(file_id: int, logger_instance: logging.Logger) -> str:
    """Retrieve the source file location URL from the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "SELECT FileLocationUrl FROM utb_ImageScraperFiles WHERE ID = ?"
        cursor.execute(query, (file_id,))
        result = cursor.fetchone()
        if not result:
            raise FileNotFoundError(f"No file location found in DB for FileID {file_id}")
        return result[0]

def get_images_excel_db(file_id: int, logger_instance: logging.Logger) -> pd.DataFrame:
    """
    Retrieve detailed image and product data for Excel processing.
    Includes all rows, even those without image results, using LEFT JOIN.
    """
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE utb_ImageScraperFiles SET CreateFileStartTime = GETDATE() WHERE ID = ?", (file_id,))
        connection.commit()
    
    # MODIFIED QUERY: Use LEFT JOIN to include all records, even those without images
    query = """
        SELECT
            s.ExcelRowID, r.ImageUrl, r.ImageUrlThumbnail, r.SortOrder,
            s.ProductBrand AS Brand, s.ProductModel AS Style,
            s.ProductColor AS Color, s.ProductCategory AS Category
        FROM utb_ImageScraperFiles f
        INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
        LEFT JOIN utb_ImageScraperResult r ON r.EntryID = s.EntryID AND r.SortOrder > 0
        WHERE f.ID = ?
        ORDER BY s.ExcelRowID, r.SortOrder
    """
    
    logger_instance.info(f"Executing query to fetch all records for FileID: {file_id}")
    return pd.read_sql_query(query, engine, params=[(file_id,)])

def update_file_location_complete(file_id: int, file_location: str, logger_instance: logging.Logger):
    """Update the completed file location URL in the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "UPDATE utb_ImageScraperFiles SET FileLocationURLComplete = ? WHERE ID = ?"
        cursor.execute(query, (file_location, file_id))
        connection.commit()
        logger_instance.info(f"Updated completed file location for FileID {file_id}")

# --- Image Downloading Functions ---
async def validate_image_response(response, url: str, logger_instance: logging.Logger) -> Tuple[bool, Optional[bytes]]:
    if response.status != 200:
        logger_instance.warning(f"Invalid status code {response.status} for URL: {url}")
        return False, None
    if not any(response.headers.get('Content-Type', '').lower().startswith(it) for it in VALID_IMAGE_TYPES):
        logger_instance.warning(f"Invalid Content-Type for URL: {url}")
        return False, None
    if int(response.headers.get('Content-Length', 0)) < MIN_IMAGE_SIZE:
        logger_instance.warning(f"Content too small for URL: {url}")
        return False, None
    return True, await response.read()

async def image_download(semaphore, item: Dict, save_path: str, session, logger_instance: logging.Logger, user_agents: List[str]):
    async with semaphore:
        row_id = item['ExcelRowID']
        image_name = str(row_id)
        loop = asyncio.get_running_loop() # Get loop here or pass it in

        for i, (url, thumb_url, sort_order) in enumerate(item['image_options']):
            
            async def attempt_download(download_url: str, is_thumb: bool) -> bool:
                if not download_url: return False
                try:
                    headers = {"User-Agent": random.choice(user_agents)}
                    async with session.get(download_url, headers=headers) as response:
                        is_valid, data = await validate_image_response(response, download_url, logger_instance)
                        if not is_valid: return False
                        
                        final_path = os.path.join(save_path, f"{image_name}.png")

                        # --- MODIFIED PART ---
                        def save_image_sync():
                            with PILImage.open(BytesIO(data)) as img:
                                img.save(final_path, 'PNG')
                        
                        await loop.run_in_executor(None, save_image_sync)
                        # --- END MODIFIED PART ---
                                
                        logger_instance.info(f"SUCCESS (SortOrder {sort_order}) for Row {row_id} from {'thumbnail' if is_thumb else 'main'} URL.")
                        return True
                except Exception as e:
                    logger_instance.warning(f"Download attempt {i+1} failed for Row {row_id} ({'thumb' if is_thumb else 'main'}): {e}")
                    return False

            if await attempt_download(url, is_thumb=False):
                return True
            if await attempt_download(thumb_url, is_thumb=True):
                return True
        
        logger_instance.error(f"All download attempts FAILED for Row {row_id}. No more images to try.")
        return False

async def download_all_images(data: List[Dict], save_path: str, logger_instance: logging.Logger, user_agents: List[str]):
    """Download all images concurrently, trying fallbacks for each item."""
    domains = []
    for item in data:
        if item.get('image_options'):
            first_url = item['image_options'][0][0]
            if first_url:
                domains.append(tldextract.extract(first_url).top_domain_under_public_suffix)
    
    pool_size = min(500, max(10, len(Counter(domains)) * 2))
    timeout_config = ClientTimeout(total=60)
    retry_options = ExponentialRetry(attempts=3, start_timeout=3)
    
    async with RetryClient(raise_for_status=False, retry_options=retry_options, timeout=timeout_config) as session:
        semaphore = asyncio.Semaphore(pool_size)
        tasks = [image_download(semaphore, item, save_path, session, logger_instance, user_agents) for item in data]
        results = await asyncio.gather(*tasks)
        
        failed_count = len([res for res in results if not res])
        if failed_count > 0:
            logger_instance.warning(f"{failed_count} product rows failed to download any image.")

# --- Image & Excel Processing Functions ---
def resize_image(image_path: str, logger_instance: logging.Logger) -> bool:
    try:
        img = PILImage.open(image_path)
        if img.mode == 'RGBA':
            background = PILImage.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3]); img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        pixels = np.array(img)
        border_pixels = np.concatenate([pixels[0, :], pixels[-1, :], pixels[:, 0], pixels[:, -1]])
        colors, counts = np.unique(border_pixels, axis=0, return_counts=True)
        most_common = colors[counts.argmax()]
        if np.sum(most_common) < 700:
            mask = np.all(np.abs(pixels - most_common) <= 15, axis=2)
            pixels[mask] = [255, 255, 255]
            img = PILImage.fromarray(pixels)

        h, w = img.size
        if h > MAX_IMAGE_DIMENSION or w > MAX_IMAGE_DIMENSION:
            img.thumbnail((MAX_IMAGE_DIMENSION, MAX_IMAGE_DIMENSION))

        img.save(image_path, 'PNG')
        return True
    except Exception as e:
        logger_instance.error(f"Error resizing image {image_path}: {e}", exc_info=True)
        return False
def get_last_non_empty_row(ws, column: str, header_row: int, logger_instance: logging.Logger) -> int:
    """Find the last non-empty row in the specified column, starting after header_row."""
    last_row = header_row
    for row in ws[f"{column}{header_row + 1}:{column}{ws.max_row}"]:
        if row[0].value is not None and str(row[0].value).strip():
            last_row = max(last_row, row[0].row)
    logger_instance.info(f"Last non-empty row in column {column}: {last_row}")
    return last_row
def verify_and_process_image(image_path: str, logger_instance: logging.Logger) -> bool:
    try:
        with PILImage.open(image_path) as img: img.verify()
        if os.path.getsize(image_path) < MAX_IMAGE_VERIFY_SIZE:
            logger_instance.warning(f"File may be too small: {image_path}"); return False
        return resize_image(image_path, logger_instance)
    except Exception:
        logger_instance.error(f"Image verification failed for {image_path}", exc_info=True); return False

def write_excel_distro(local_filename: str, temp_dir: str, image_data: List[Dict], header_row: int, logger_instance: logging.Logger):
    wb = load_workbook(local_filename)
    ws = wb.active
    image_map = {int(Path(f).stem): f for f in os.listdir(temp_dir) if Path(f).stem.isdigit()}

    # Get default row height from the template
    DEFAULT_ROW_HEIGHT_POINTS = ws.row_dimensions.get(header_row + 1, {}).height
    if DEFAULT_ROW_HEIGHT_POINTS is None:
        DEFAULT_ROW_HEIGHT_POINTS = 12.75  # Excel default height in points
        logger_instance.info(f"No row height set in template for row {header_row + 1}, using default {DEFAULT_ROW_HEIGHT_POINTS} points")
    else:
        logger_instance.info(f"Using template row height: {DEFAULT_ROW_HEIGHT_POINTS} points from row {header_row + 1}")

    # Create a mapping of ExcelRowID to metadata for quick lookup
    row_data_map = {item['ExcelRowID']: item for item in image_data}

    # NEW: Call get_last_non_empty_row to find the last non-empty row in column B
    last_non_empty_row = get_last_non_empty_row(ws, column='B', header_row=header_row, logger_instance=logger_instance)

    # Determine the range of ExcelRowIDs to process
    if image_data:
        min_row_id = min(item['ExcelRowID'] for item in image_data)
        max_row_id = max(item['ExcelRowID'] for item in image_data)
        # Adjust max_row_id to account for the last non-empty row in the template
        max_row_id = max(max_row_id, last_non_empty_row - header_row)
        logger_instance.info(f"Row range: ExcelRowID {min_row_id} to {max_row_id}, adjusted for template last row {last_non_empty_row}")
    else:
        min_row_id = 1
        max_row_id = last_non_empty_row - header_row if last_non_empty_row > header_row else 1
        logger_instance.warning(f"No data in image_data, setting range based on template last row {last_non_empty_row}")

    # Ensure enough rows exist in the worksheet
    max_needed_row = max_row_id + header_row
    if ws.max_row < max_needed_row:
        logger_instance.info(f"Appending {max_needed_row - ws.max_row} rows to worksheet")
        for row_num in range(ws.max_row + 1, max_needed_row + 1):
            ws.append([''] * ws.max_column)  # Append empty row to match column count
            ws.row_dimensions[row_num].height = DEFAULT_ROW_HEIGHT_POINTS

    # Process all rows in the expected range to avoid gaps
    for row_id in range(min_row_id, max_row_id + 1):
        row_num = row_id + header_row
        ws.row_dimensions[row_num].height = DEFAULT_ROW_HEIGHT_POINTS

        if row_id in row_data_map:
            item = row_data_map[row_id]
            # Write image if available
            if row_id in image_map:
                image_path = os.path.join(temp_dir, image_map[row_id])
                if verify_and_process_image(image_path, logger_instance):
                    img = Image(image_path)
                    img.anchor = f"A{row_num}"
                    ws.add_image(img)
                    # Adjust row height based on image if needed
                    img_height_pixels = img.height if hasattr(img, 'height') else 0
                    img_height_points = img_height_pixels * 72 / 96  # Assuming 96 DPI
                    ws.row_dimensions[row_num].height = max(DEFAULT_ROW_HEIGHT_POINTS, img_height_points)
                    logger_instance.info(f"Added image for Row {row_id} at Excel row {row_num}, height set to {ws.row_dimensions[row_num].height} points")
                else:
                    logger_instance.warning(f"Image processing failed for Row {row_id}, writing metadata only")
            else:
                logger_instance.info(f"No image found for Row {row_id}, writing metadata only")

            # Write metadata
            ws[f"B{row_num}"] = item.get('Brand', '')
            ws[f"D{row_num}"] = item.get('Style', '')
            ws[f"E{row_num}"] = item.get('Color', '')
            ws[f"H{row_num}"] = item.get('Category', '')
            logger_instance.info(f"Wrote metadata for Row {row_id} at Excel row {row_num}")
        else:
            # Fill missing row with empty metadata
            ws[f"B{row_num}"] = ''
            ws[f"D{row_num}"] = ''
            ws[f"E{row_num}"] = ''
            ws[f"H{row_num}"] = ''
            logger_instance.info(f"Filled missing row {row_num} (ExcelRowID {row_id}) with empty metadata")
        # Remove rows after the last data row
        if ws.max_row > max_row_id + header_row:
            logger_instance.info(f"Deleting {ws.max_row - (max_row_id + header_row)} rows after row {max_row_id + header_row}")
            ws.delete_rows(max_row_id + header_row + 1, ws.max_row - (max_row_id + header_row))
        logger_instance.info("Setting worksheet view to A1.")
        ws.sheet_view.topLeftCell = 'A1'
        ws.active_cell = 'A1'
        wb.save(local_filename)
        logger_instance.info(f"Excel file saved: {local_filename}")

def write_excel_generic(local_filename: str, temp_dir: str, header_row: int, row_offset: int, logger_instance: logging.Logger):
    try:
        wb = load_workbook(local_filename); ws = wb.active
        image_map = {int(Path(f).stem): f for f in os.listdir(temp_dir) if Path(f).stem.isdigit()}

        for row_id, image_file in image_map.items():
            image_path = os.path.join(temp_dir, image_file)
            if verify_and_process_image(image_path, logger_instance):
                img = Image(image_path)
                adjusted_row = row_id + header_row + row_offset
                img.anchor = f"A{adjusted_row}"; ws.add_image(img)
        logger_instance.info("Setting worksheet view to A1.")
        ws.sheet_view.topLeftCell = 'A1'
        ws.active_cell = 'A1'
        wb.save(local_filename)
    except Exception as e:
        logger_instance.error(f"Error writing to generic Excel file: {e}", exc_info=True)
        raise

def find_header_row_index(excel_file: str, logger_instance: logging.Logger) -> Optional[int]:
    try:
        wb = load_workbook(excel_file, read_only=True); ws = wb.active
        keywords = {'id', 'name', 'product', 'brand', 'sku', 'category', 'color', 'price', 'description'}
        best_row, best_score = None, 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if not row: continue
            cells = [str(c).strip().lower() for c in row if isinstance(c, str)]
            score = len(cells) + len(set(cells)) + sum(any(k in c for k in keywords) for c in cells)
            if score > best_score: best_score, best_row = score, i
        if best_row and best_score > 3: return best_row
        return None
    except Exception as e:
        logger_instance.error(f"Could not find header row: {e}"); return None

# --- Main Background Task ---
async def generate_download_file(file_id: str, row_offset: int = 0):
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    logger_instance = setup_logging(file_id, timestamp)
    temp_images_dir, temp_excel_dir = "", ""
    file_id_int = int(file_id)

    try:
        user_agents = get_user_agents(logger_instance)
        temp_images_dir, temp_excel_dir = await create_temp_dirs(file_id)
        
        logger_instance.info("Fetching all image data from database...")
        images_df = get_images_excel_db(file_id_int, logger_instance)
        
        # NEW: Fetch all records if images_df is empty to ensure metadata is processed
        if images_df.empty:
            logger_instance.warning(f"No image data found for FileID {file_id}. Fetching all records for metadata.")
            query = """
                SELECT
                    s.ExcelRowID,
                    s.ProductBrand AS Brand, s.ProductModel AS Style,
                    s.ProductColor AS Color, s.ProductCategory AS Category
                FROM utb_ImageScraperFiles f
                INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
                WHERE f.ID = ?
                ORDER BY s.ExcelRowID
            """
            images_df = pd.read_sql_query(query, engine, params=[(file_id,)])
            if images_df.empty:
                logger_instance.error(f"No records found for FileID {file_id}. The job cannot proceed.")
                return

        logger_instance.info("Grouping data by product to prepare for download attempts...")
        grouped_data = []
        for _, group in images_df.groupby('ExcelRowID'):
            first_row = group.iloc[0]
            # Include image options only if they exist
            image_options = list(zip(group['ImageUrl'], group['ImageUrlThumbnail'], group['SortOrder'])) if 'ImageUrl' in group.columns else []
            
            grouped_data.append({
                'ExcelRowID': int(first_row['ExcelRowID']),
                'Brand': first_row['Brand'],
                'Style': first_row['Style'],
                'Color': first_row['Color'],
                'Category': first_row['Category'],
                'image_options': image_options
            })
        
        logger_instance.info(f"Data prepared for {len(grouped_data)} unique products. Starting image downloads.")
        if any(item['image_options'] for item in grouped_data):
            await download_all_images(grouped_data, temp_images_dir, logger_instance, user_agents)
        else:
            logger_instance.info("No images to download, proceeding with metadata only.")
        
        file_type_id = get_file_type_id(file_id_int, logger_instance)
        FILE_TYPE_DISTRO = 3

        if file_type_id == FILE_TYPE_DISTRO:
            logger_instance.info("Starting DISTRO file generation.")
            template_url = "https://iconluxury.group/public/ICON_DISTRO_USD_20250617.xlsx"
            file_name = os.path.basename(urllib.parse.unquote(template_url))
            local_filename = os.path.join(temp_excel_dir, file_name)
            header_row = 5

            res = requests.get(template_url, timeout=60); res.raise_for_status()
            with open(local_filename, "wb") as f: f.write(res.content)
            
            write_excel_distro(local_filename, temp_images_dir, grouped_data, header_row, logger_instance)
        else:
            logger_instance.info("Starting GENERIC file generation.")
            file_url = get_file_location(file_id_int, logger_instance)
            file_name = os.path.basename(urllib.parse.unquote(file_url))
            local_filename = os.path.join(temp_excel_dir, file_name)

            res = requests.get(file_url, timeout=60); res.raise_for_status()
            with open(local_filename, "wb") as f: f.write(res.content)
            
            header_row = find_header_row_index(local_filename, logger_instance) or 0
            write_excel_generic(local_filename, temp_images_dir, header_row, row_offset, logger_instance)

        processed_file_name = f"{Path(file_name).stem}_processed_{timestamp}.xlsx"
        public_url = await upload_file_to_space(local_filename, save_as=f"processed_files/{processed_file_name}", file_id=file_id_int, is_public=True)
        update_file_location_complete(file_id_int, public_url, logger_instance)
        await send_email(to_emails='nik@iconluxurygroup.com', subject=f'File Processed: {file_name}', download_url=public_url, job_id=file_id)
        
        logger_instance.info(f"Successfully completed job for FileID {file_id}.")

    except Exception as e:
        logger_instance.error(f"FATAL ERROR for FileID {file_id}: {e}", exc_info=True)
    finally:
        if temp_images_dir and temp_excel_dir:
            await cleanup_temp_dirs([temp_images_dir, temp_excel_dir])

# --- FastAPI Endpoints ---
@app.post("/generate-download-file/")
async def process_file(background_tasks: BackgroundTasks, file_id: int, row_offset: Optional[int] = 0):
    logger.info(f"Received request for FileID: {file_id} with row_offset={row_offset}")
    background_tasks.add_task(generate_download_file, str(file_id), row_offset)
    return {"message": "Processing started. You will be notified upon completion."}

@app.get("/")
def read_root():
    return {"message": "Image Scraper Service is running."}

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Uvicorn server on http://0.0.0.0:8080")
    uvicorn.run(app, host="0.0.0.0", port=8080)